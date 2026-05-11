import os
import sys

# macOS：FAISS 与 PyTorch 同时链接不同 OpenMP/BLAS 时易 SIGSEGV；需在 numpy/torch 初始化前收紧线程
if sys.platform == "darwin":
    for _k, _v in (
        ("OMP_NUM_THREADS", "1"),
        ("MKL_NUM_THREADS", "1"),
        ("VECLIB_MAXIMUM_THREADS", "1"),
        ("NUMEXPR_MAX_THREADS", "1"),
        ("OPENBLAS_NUM_THREADS", "1"),
    ):
        os.environ.setdefault(_k, _v)

from flask import Flask, render_template, request, jsonify, send_file, send_from_directory
from data_mgr import DataManager
from license_utils import LicenseManager
import signal
import faulthandler
import shutil
import time
import threading
import traceback
import io
import json
import uuid
import zipfile
import subprocess
import platform
import datetime
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook
import utils

_single_instance_lock_fh = None

if hasattr(signal, 'SIGUSR1'):
    faulthandler.register(signal.SIGUSR1, all_threads=True, chain=False)


def _resolve_app_paths():
    """
    PyInstaller 打包后：只读资源在 sys._MEIPASS；数据库/上传/缓存必须写在 exe 旁可写目录，
    否则写入 _MEIPASS 会失败或无法持久化。
    """
    if getattr(sys, 'frozen', False):
        resource_root = getattr(sys, '_MEIPASS', os.path.dirname(sys.executable))
        exe_dir = os.path.dirname(sys.executable)
        # macOS .app：sys.executable 通常在 .../ProImage_AI.app/Contents/MacOS/
        # 交付时更希望把可写数据放在 .app 同级目录：<folder>/ProImage_data
        bundle_dir = None
        try:
            p = exe_dir
            if p.endswith(os.path.join('Contents', 'MacOS')):
                bundle_dir = os.path.dirname(os.path.dirname(p))  # .../ProImage_AI.app
        except Exception:
            bundle_dir = None

        external_data_root = None
        if bundle_dir:
            external_data_root = os.path.join(os.path.dirname(bundle_dir), 'ProImage_data')

        candidate_roots = [r for r in [external_data_root, os.path.join(exe_dir, 'ProImage_data')] if r]
        data_root = None
        for r in candidate_roots:
            try:
                os.makedirs(r, exist_ok=True)
                data_root = r
                break
            except Exception:
                continue
        if not data_root:
            data_root = os.path.join(exe_dir, 'ProImage_data')
            os.makedirs(data_root, exist_ok=True)
        os.makedirs(data_root, exist_ok=True)
        os.makedirs(os.path.join(data_root, 'uploads'), exist_ok=True)
        os.makedirs(os.path.join(data_root, 'img'), exist_ok=True)
    else:
        resource_root = os.path.dirname(os.path.abspath(__file__))
        data_root = resource_root
    return resource_root, data_root


resource_root, data_root = _resolve_app_paths()
# 类目规则页默认类目表（与「下载模板」同源时可复制到 data/default_meituan_categories.xlsx）
DEFAULT_RULE_CATEGORIES_XLSX = os.path.join(resource_root, "data", "default_meituan_categories.xlsx")
CATEGORY_L1_BUCKET_TAGS_JSON = os.path.join(resource_root, "data", "category_l1_bucket_tags.json")
# 冻结版：分析线程里相对路径 img/、query_img/ 与 DataManager 使用同一根目录
if getattr(sys, 'frozen', False):
    os.chdir(data_root)


def _acquire_single_instance_lock():
    """
    防止 macOS 启动很慢时被多次双击，导致同时启动多个实例。
    仅在 frozen（.app）模式启用。
    """
    global _single_instance_lock_fh
    if not getattr(sys, 'frozen', False):
        return True

    try:
        import fcntl  # macOS/Linux
    except Exception:
        return True

    lock_path = os.path.join(data_root, "ProImage_AI.lock")
    try:
        fh = open(lock_path, "w")
        fcntl.flock(fh.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
        fh.write(str(os.getpid()))
        fh.flush()
        _single_instance_lock_fh = fh
        return True
    except Exception:
        try:
            fh.close()
        except Exception:
            pass
        return False


if not _acquire_single_instance_lock():
    raise SystemExit(0)

_template = os.path.join(resource_root, 'templates')
_static = os.path.join(resource_root, 'static')
if os.path.isdir(_static):
    app = Flask(__name__, template_folder=_template, static_folder=_static, static_url_path='/static')
else:
    app = Flask(__name__, template_folder=_template)

app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB
dm = DataManager(data_root)

# 放在单实例锁之后，避免多次双击时重复触发重依赖初始化
import extract_info_ai2  # noqa: E402
import main_030822  # noqa: E402

# ── Analysis progress tracking ──
_analysis_progress = {}
_progress_lock = threading.Lock()
_ops_tasks = {}
_ops_lock = threading.Lock()

def _init_progress(pid, use_ai, main_name, comp_names):
    steps = []
    if use_ai:
        steps.append({"label": f"AI提取 {main_name}", "status": "pending", "detail": ""})
        for cn in comp_names:
            steps.append({"label": f"AI提取 {cn}", "status": "pending", "detail": ""})
    for cn in comp_names:
        steps.append({"label": f"向量分析 {cn}", "status": "pending", "detail": ""})
    steps.append({"label": f"查询匹配 {main_name}", "status": "pending", "detail": ""})
    prog = {"started_at": time.time(), "steps": steps}
    with _progress_lock:
        _analysis_progress[pid] = prog
    return prog

def _update_step(pid, step_idx, status, detail=""):
    with _progress_lock:
        prog = _analysis_progress.get(pid)
        if not prog or step_idx >= len(prog["steps"]):
            return
        step = prog["steps"][step_idx]
        step["status"] = status
        step["detail"] = detail
        if status == "running" and not step.get("started_at"):
            step["started_at"] = time.time()
        if status == "done" and not step.get("ended_at"):
            step["ended_at"] = time.time()

def _clear_progress(pid):
    with _progress_lock:
        _analysis_progress.pop(pid, None)


# 成功/失败后延迟清理内存进度，让前端能至少轮询到一次「全步完成 / 100%」再消失
_ANALYSIS_PROGRESS_HOLD_S = 5.0


def _schedule_clear_progress(pid):
    threading.Timer(_ANALYSIS_PROGRESS_HOLD_S, lambda p=pid: _clear_progress(p)).start()

MAX_FILE_SIZE = 80 * 1024 * 1024  # 80MB per file
ALLOWED_EXTENSIONS = {'.xlsx', '.xls'}

def _validate_upload(file_storage, label):
    """Validate file extension and size. Returns error message or None."""
    ext = os.path.splitext(file_storage.filename)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        return f"{label}：不支持的文件格式 ({ext})，仅支持 .xlsx / .xls"
    file_storage.seek(0, 2)
    size = file_storage.tell()
    file_storage.seek(0)
    if size > MAX_FILE_SIZE:
        return f"{label}：文件过大 ({size // 1024 // 1024}MB)，上限 {MAX_FILE_SIZE // 1024 // 1024}MB"
    return None


def _safe_upload_filename(filename, fallback):
    base = secure_filename(filename or "")
    orig_ext = os.path.splitext(filename or "")[1].lower()
    if not base:
        base = fallback
    elif orig_ext and not os.path.splitext(base)[1]:
        base = base + orig_ext
    return base


def _norm_cell(v):
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in ("nan", "none", "null") else s


def _build_category_tree(rows):
    l1_aliases = ("美团一级类目", "一级类目", "美团类目一级", "美团1级类目")
    l2_aliases = ("美团二级类目", "二级类目", "美团类目二级", "美团2级类目")
    l3_aliases = ("美团三级类目", "三级类目", "美团类目三级", "美团3级类目")

    def first_val(item, aliases):
        for key in aliases:
            if key in item:
                v = _norm_cell(item.get(key))
                if v:
                    return v
        return ""

    l1_map = {}
    for row in rows:
        l3 = first_val(row, l3_aliases)
        if not l3:
            continue
        l1 = first_val(row, l1_aliases) or "未分类一级类目"
        l2 = first_val(row, l2_aliases) or "未分类二级类目"
        l1_entry = l1_map.setdefault(l1, {"name": l1, "children_map": {}, "l3_count": 0})
        l2_entry = l1_entry["children_map"].setdefault(l2, {"name": l2, "children": [], "seen": set()})
        if l3 not in l2_entry["seen"]:
            l2_entry["children"].append({"name": l3})
            l2_entry["seen"].add(l3)
            l1_entry["l3_count"] += 1

    items = []
    for l1_name in sorted(l1_map.keys()):
        l1_entry = l1_map[l1_name]
        children = []
        for l2_name in sorted(l1_entry["children_map"].keys()):
            l2_entry = l1_entry["children_map"][l2_name]
            children.append({
                "name": l2_name,
                "l3_count": len(l2_entry["children"]),
                "children": sorted(l2_entry["children"], key=lambda x: x["name"]),
            })
        items.append({
            "name": l1_name,
            "l3_count": l1_entry["l3_count"],
            "children": children,
        })

    return {
        "items": items,
        "l1_count": len(items),
        "l3_count": sum(item["l3_count"] for item in items),
    }


def _workbook_to_rows(wb):
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return []
    out = []
    for row in rows:
        if all(cell is None for cell in row):
            continue
        out.append(dict(zip(headers, row)))
    return out


def _excel_file_to_rows(file_storage):
    wb = load_workbook(file_storage, data_only=True)
    return _workbook_to_rows(wb)


def _excel_path_to_rows(path: str):
    wb = load_workbook(path, data_only=True)
    return _workbook_to_rows(wb)


def _ops_now():
    return time.strftime("%Y-%m-%d %H:%M:%S")


def _ops_task_dir(task_id):
    return os.path.join(data_root, "uploads", "ops_tasks", task_id)


def _ops_public_task(task):
    return {
        "task_id": task.get("task_id"),
        "kind": task.get("kind"),
        "status": task.get("status"),
        "message": task.get("message", ""),
        "error": task.get("error", ""),
        "created_at": task.get("created_at"),
        "started_at": task.get("started_at"),
        "ended_at": task.get("ended_at"),
        "steps": task.get("steps", []),
        "download_ready": bool(task.get("result_path") and os.path.exists(task.get("result_path", ""))),
        "download_name": task.get("download_name", ""),
        "result_kind": task.get("result_kind", ""),
        "source_task_id": task.get("source_task_id", ""),
    }


def _ops_create_task(kind, steps, message=""):
    task_id = uuid.uuid4().hex[:12]
    task = {
        "task_id": task_id,
        "kind": kind,
        "status": "pending",
        "message": message,
        "error": "",
        "created_at": _ops_now(),
        "started_at": None,
        "ended_at": None,
        "steps": [{"label": s, "status": "pending", "detail": ""} for s in steps],
        "result_path": "",
        "download_name": "",
        "result_kind": "",
    }
    with _ops_lock:
        _ops_tasks[task_id] = task
    return task


def _ops_get_task(task_id):
    with _ops_lock:
        return _ops_tasks.get(task_id)


def _ops_set_task(task_id, **kwargs):
    with _ops_lock:
        task = _ops_tasks.get(task_id)
        if not task:
            return
        task.update(kwargs)


def _ops_update_step(task_id, idx, status, detail=""):
    with _ops_lock:
        task = _ops_tasks.get(task_id)
        if not task or idx < 0 or idx >= len(task.get("steps", [])):
            return
        step = task["steps"][idx]
        step["status"] = status
        step["detail"] = detail or ""
        now = time.time()
        if status == "running" and not step.get("started_at"):
            step["started_at"] = now
        if status in ("done", "failed") and not step.get("ended_at"):
            step["ended_at"] = now


def _ops_fail_task(task_id, err):
    _ops_set_task(task_id, status="failed", error=str(err), ended_at=_ops_now())


def _ops_file_label(file_storage, fallback):
    name = (getattr(file_storage, "filename", "") or "").strip()
    return name or fallback


def _ops_safe_filename(filename, fallback):
    ext = os.path.splitext(filename or "")[1].lower() or ".xlsx"
    safe = _safe_upload_filename(filename, fallback)
    if not os.path.splitext(safe)[1]:
        safe += ext
    return safe


def _ops_save_file(file_storage, dest_dir, prefix, idx=0):
    os.makedirs(dest_dir, exist_ok=True)
    original = _ops_file_label(file_storage, f"{prefix}_{idx}.xlsx")
    safe = _ops_safe_filename(original, f"{prefix}_{idx}.xlsx")
    base, ext = os.path.splitext(safe)
    filename = f"{prefix}_{idx}__{base}{ext}"
    path = os.path.join(dest_dir, filename)
    n = 1
    while os.path.exists(path):
        filename = f"{prefix}_{idx}__{base}_{n}{ext}"
        path = os.path.join(dest_dir, filename)
        n += 1
    file_storage.save(path)
    return {"path": path, "original_name": original, "safe_name": filename}


def _ops_validate_excel_uploads(main_file, comp_files):
    if not main_file or not main_file.filename:
        return "请上传主店文件"
    err = _validate_upload(main_file, "主店文件")
    if err:
        return err
    valid_comp_files = [f for f in comp_files if f and f.filename]
    if not valid_comp_files:
        return "请至少上传一个竞店文件"
    for f in valid_comp_files:
        err = _validate_upload(f, f"竞店文件 ({f.filename})")
        if err:
            return err
    return None


def _ops_validate_astar_input_columns(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        headers = {str(c).strip() for c in (row or []) if c is not None}
        if "商品名称" not in headers or not ({"规格", "规格名称"} & headers):
            raise ValueError("缺少必需列：商品名称 + 规格/规格名称")
    finally:
        wb.close()


def _ops_copy_to_dir(src, dest_dir, prefix, idx, original_name=None):
    os.makedirs(dest_dir, exist_ok=True)
    original = original_name or os.path.basename(src)
    safe = _ops_safe_filename(original, f"{prefix}_{idx}.xlsx")
    base, ext = os.path.splitext(safe)
    filename = f"{prefix}_{idx}__{base}{ext}"
    dest = os.path.join(dest_dir, filename)
    n = 1
    while os.path.exists(dest):
        filename = f"{prefix}_{idx}__{base}_{n}{ext}"
        dest = os.path.join(dest_dir, filename)
        n += 1
    shutil.copy2(src, dest)
    return {"path": dest, "original_name": original, "safe_name": filename}


def _ops_zip_files(files, zip_path):
    os.makedirs(os.path.dirname(zip_path), exist_ok=True)
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        used = set()
        for item in files:
            arc = item.get("arcname") or item.get("original_name") or os.path.basename(item["path"])
            arc = arc.replace("/", "_").replace("\\", "_")
            if arc in used:
                base, ext = os.path.splitext(arc)
                n = 1
                while f"{base}_{n}{ext}" in used:
                    n += 1
                arc = f"{base}_{n}{ext}"
            used.add(arc)
            zf.write(item["path"], arcname=arc)


def _ops_rule_template_from_request(raw_rule_id):
    templates = dm.list_rule_templates()
    tid = None
    try:
        tid = int(raw_rule_id) if str(raw_rule_id or "").strip() else None
    except ValueError:
        tid = None
    if not tid:
        prod = next((t for t in templates if t.get("name") == "生产规则V1"), None)
        tid = int(prod["id"]) if prod else (int(templates[0]["id"]) if templates else None)
    t = dm.get_rule_template(tid) if tid else None
    if not t:
        raise ValueError("未找到可用的类目规则模板")
    return t


def _ops_load_private_key(uploaded_key=None):
    from cryptography.hazmat.primitives import serialization

    key_bytes = None
    if uploaded_key and uploaded_key.filename:
        key_bytes = uploaded_key.read()
        uploaded_key.seek(0)
    else:
        for p in (
            os.path.join(resource_root, "vendor", "private_key.pem"),
            os.path.join(data_root, "vendor", "private_key.pem"),
            os.path.join(resource_root, "private_key.pem"),
            os.path.join(data_root, "private_key.pem"),
        ):
            if os.path.isfile(p):
                with open(p, "rb") as f:
                    key_bytes = f.read()
                break
    if not key_bytes:
        raise ValueError("未找到 private_key.pem；请上传私钥文件，或放到 vendor/private_key.pem")
    return serialization.load_pem_private_key(key_bytes, password=None)


def _ops_default_private_key_status():
    candidates = (
        (os.path.join(resource_root, "vendor", "private_key.pem"), "vendor/private_key.pem"),
        (os.path.join(data_root, "vendor", "private_key.pem"), "vendor/private_key.pem"),
        (os.path.join(resource_root, "private_key.pem"), "private_key.pem"),
        (os.path.join(data_root, "private_key.pem"), "private_key.pem"),
    )
    for path, label in candidates:
        if os.path.isfile(path):
            return {"configured": True, "path_label": label}
    return {"configured": False, "path_label": ""}


def _ops_create_license_file(hwids, expires_days, out_path, uploaded_key=None):
    from cryptography.hazmat.primitives import hashes
    from cryptography.hazmat.primitives.asymmetric import padding
    import base64

    private_key = _ops_load_private_key(uploaded_key)
    expires = (datetime.datetime.now() + datetime.timedelta(days=int(expires_days))).strftime("%Y-%m-%d")
    data = {"hwids": hwids, "expires": expires, "version": "1.0"}
    data_json = json.dumps(data, ensure_ascii=False)
    data_b64 = base64.b64encode(data_json.encode()).decode()
    signature = private_key.sign(
        data_json.encode(),
        padding.PSS(mgf=padding.MGF1(hashes.SHA256()), salt_length=padding.PSS.MAX_LENGTH),
        hashes.SHA256(),
    )
    sig_b64 = base64.b64encode(signature).decode()
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(f"{data_b64}.{sig_b64}")
    return expires


_OPS_PYARMOR_FILES = [
    "app.py",
    "data_mgr.py",
    "data_mgr_base.py",
    "data_mgr_import.py",
    "data_mgr_query.py",
    "data_mgr_ops.py",
    "data_mgr_export.py",
    "data_mgr_rule_templates.py",
    "license_utils.py",
    "main_030822.py",
    "extract_info_ai2.py",
    "product_text_extract.py",
    "post_match_engine.py",
    "utils.py",
    "merge_sku_data.py",
]


def _ops_pyarmor_command():
    pyarmor = shutil.which("pyarmor")
    if pyarmor:
        return [pyarmor]
    user_bin = os.path.join(os.path.expanduser("~"), "Library", f"Python/{sys.version_info.major}.{sys.version_info.minor}", "bin", "pyarmor")
    if os.path.isfile(user_bin):
        return [user_bin]
    return [sys.executable, "-m", "pyarmor"]


def _ops_verify_pyarmor_output(obf_dir):
    missing = [name for name in _OPS_PYARMOR_FILES if not os.path.isfile(os.path.join(obf_dir, name))]
    runtime_dirs = [
        name for name in os.listdir(obf_dir) if name.startswith("pyarmor_runtime_") and os.path.isdir(os.path.join(obf_dir, name))
    ] if os.path.isdir(obf_dir) else []
    if missing or not runtime_dirs:
        parts = []
        if missing:
            parts.append("缺少文件：" + ", ".join(missing))
        if not runtime_dirs:
            parts.append("缺少 pyarmor_runtime_* 目录")
        raise RuntimeError("PyArmor 混淆结果不完整；" + "；".join(parts))


def _ops_run_command(task_id, step_idx, cmd, cwd):
    _ops_update_step(task_id, step_idx, "running", " ".join(cmd))
    proc = subprocess.Popen(
        cmd,
        cwd=cwd,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        bufsize=1,
    )
    last = ""
    if proc.stdout:
        for line in proc.stdout:
            line = line.rstrip()
            if not line:
                continue
            last = line[-240:]
            _ops_update_step(task_id, step_idx, "running", last)
    code = proc.wait()
    if code != 0:
        raise RuntimeError(f"命令失败({code}): {' '.join(cmd)}\n{last}")
    _ops_update_step(task_id, step_idx, "done", "完成")


def _ops_zip_path(src_path, zip_path):
    os.makedirs(os.path.dirname(zip_path), exist_ok=True)
    if platform.system() == "Darwin" and os.path.isdir(src_path) and src_path.endswith(".app"):
        subprocess.run(
            ["ditto", "-c", "-k", "--sequesterRsrc", "--keepParent", src_path, zip_path],
            check=True,
        )
        return zip_path
    base = zip_path[:-4] if zip_path.lower().endswith(".zip") else zip_path
    if os.path.isdir(src_path):
        shutil.make_archive(base, "zip", root_dir=os.path.dirname(src_path), base_dir=os.path.basename(src_path))
    else:
        _ops_zip_files([{"path": src_path, "arcname": os.path.basename(src_path)}], zip_path)
    return zip_path if zip_path.lower().endswith(".zip") else base + ".zip"


def _ops_license_error_response():
    is_valid, msg = check_license()
    if is_valid:
        return None
    return jsonify({"status": "error", "message": msg or "授权未通过"}), 403

# --- License Check Logic ---
LICENSE_FILE = os.path.join(data_root, "license.dat")
CURRENT_HWID = LicenseManager.get_hwid()

def check_license():
    if not os.path.exists(LICENSE_FILE): return False, "License file missing"
    with open(LICENSE_FILE, "r") as f: content = f.read().strip()
    return LicenseManager.verify_license(content, CURRENT_HWID)

def get_license_details():
    if not os.path.exists(LICENSE_FILE):
        return {
            "valid": False,
            "message": "License file missing",
            "expires": None,
            "days_remaining": None,
        }
    with open(LICENSE_FILE, "r") as f:
        content = f.read().strip()
    return LicenseManager.verify_license_detailed(content, CURRENT_HWID)

@app.errorhandler(413)
def request_entity_too_large(e):
    return jsonify({"status": "error", "message": "上传文件总大小超过 100MB 限制"}), 413

@app.route('/api/license_info')
def get_license_info():
    d = get_license_details()
    return jsonify({
        "hwid": CURRENT_HWID,
        "is_valid": d["valid"],
        "message": d["message"],
        "expires": d.get("expires"),
        "days_remaining": d.get("days_remaining"),
    })

@app.route('/')
def projects_page():
    is_valid, _ = check_license()
    if not is_valid: return render_template('activate.html', hwid=CURRENT_HWID)
    return render_template('projects.html')

@app.route('/dashboard')
def index():
    is_valid, _ = check_license()
    if not is_valid: return render_template('activate.html', hwid=CURRENT_HWID)
    return render_template('index.html', active_project=dm.active_project_name)


@app.route('/ops-tools')
def ops_tools_page():
    is_valid, _ = check_license()
    if not is_valid:
        return render_template('activate.html', hwid=CURRENT_HWID)
    return render_template('ops_tools.html')


@app.route('/api/ops/astar-extract', methods=['POST'])
def api_ops_astar_extract():
    license_err = _ops_license_error_response()
    if license_err:
        return license_err
    main_file = request.files.get('main_file')
    comp_files = [f for f in request.files.getlist('comp_files') if f and f.filename]
    err = _ops_validate_excel_uploads(main_file, comp_files)
    if err:
        return jsonify({"status": "error", "message": err}), 400

    api_key = (request.form.get("api_key") or "").strip()
    if not api_key:
        return jsonify({"status": "error", "message": "请填写 Gemini API Key"}), 400
    ai_model_name = (request.form.get("ai_model_name") or "").strip()
    kimi_api_key = (request.form.get("kimi_api_key") or "").strip()
    kimi_model_name = (request.form.get("kimi_model_name") or "").strip()

    labels = [_ops_file_label(main_file, "主店文件")] + [_ops_file_label(f, f"竞店{i+1}") for i, f in enumerate(comp_files)]
    task = _ops_create_task("astar", [f"A* 提取 {label}" for label in labels], "A* 提取排队中")
    task_id = task["task_id"]
    task_dir = _ops_task_dir(task_id)
    sources_dir = os.path.join(task_dir, "sources")
    astar_dir = os.path.join(task_dir, "astar")
    os.makedirs(sources_dir, exist_ok=True)
    os.makedirs(astar_dir, exist_ok=True)

    saved_main = _ops_save_file(main_file, sources_dir, "main", 0)
    saved_comps = [_ops_save_file(f, sources_dir, "comp", i) for i, f in enumerate(comp_files)]
    astar_main = _ops_copy_to_dir(saved_main["path"], astar_dir, "main", 0, saved_main["original_name"])
    astar_comps = [
        _ops_copy_to_dir(item["path"], astar_dir, "comp", i, item["original_name"])
        for i, item in enumerate(saved_comps)
    ]

    def _run_astar_bg():
        _ops_set_task(task_id, status="running", started_at=_ops_now(), message="A* 提取中")
        try:
            items = [astar_main] + astar_comps
            for idx, item in enumerate(items):
                _ops_update_step(task_id, idx, "running", "检查表头")
                _ops_validate_astar_input_columns(item["path"])

                def _ai_cb(batch, total, _idx=idx):
                    _ops_update_step(task_id, _idx, "running", f"batch {batch}/{total}")

                extract_info_ai2.process_file_ai(
                    item["path"],
                    api_key,
                    progress_cb=_ai_cb,
                    model_name=ai_model_name,
                    fallback_api_key=kimi_api_key or None,
                    fallback_model=kimi_model_name or None,
                )
                _ops_update_step(task_id, idx, "done", "完成")

            zip_name = f"A星提取结果_{time.strftime('%Y%m%d_%H%M%S')}.zip"
            zip_path = os.path.join(task_dir, zip_name)
            zip_items = []
            zip_items.append({"path": astar_main["path"], "arcname": astar_main["original_name"]})
            for item in astar_comps:
                zip_items.append({"path": item["path"], "arcname": item["original_name"]})
            _ops_zip_files(zip_items, zip_path)
            _ops_set_task(
                task_id,
                status="done",
                ended_at=_ops_now(),
                message="A* 提取完成",
                result_path=zip_path,
                result_kind="astar_zip",
                download_name=zip_name,
                astar_main=astar_main,
                astar_comps=astar_comps,
            )
        except BaseException as e:
            traceback.print_exc()
            _ops_update_step(task_id, next((i for i, s in enumerate(_ops_get_task(task_id).get("steps", [])) if s["status"] == "running"), 0), "failed", str(e))
            _ops_fail_task(task_id, e)

    threading.Thread(target=_run_astar_bg, daemon=True).start()
    return jsonify({"status": "ok", "task_id": task_id})


@app.route('/api/ops/output-generate', methods=['POST'])
def api_ops_output_generate():
    license_err = _ops_license_error_response()
    if license_err:
        return license_err
    astar_task_id = (request.form.get("astar_task_id") or "").strip()
    use_astar_task = (request.form.get("use_astar_task") or "").strip() == "1"
    main_file = request.files.get('main_file')
    comp_files = [f for f in request.files.getlist('comp_files') if f and f.filename]

    try:
        rule_template = _ops_rule_template_from_request(request.form.get("rule_template_id"))
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 400

    from_task = None
    if use_astar_task and astar_task_id:
        from_task = _ops_get_task(astar_task_id)
        if not from_task or from_task.get("status") != "done":
            return jsonify({"status": "error", "message": "上一步 A* 任务不存在或尚未完成"}), 400
        if not from_task.get("astar_main") or not from_task.get("astar_comps"):
            return jsonify({"status": "error", "message": "上一步任务没有可用的 A* 文件"}), 400
    else:
        err = _ops_validate_excel_uploads(main_file, comp_files)
        if err:
            return jsonify({"status": "error", "message": err}), 400

    task = _ops_create_task(
        "output",
        ["准备文件"] + [f"向量分析 竞店{i+1}" for i in range(len(from_task.get("astar_comps", [])) if from_task else len(comp_files))] + ["查询匹配主店"],
        "Output 生成排队中",
    )
    task_id = task["task_id"]
    task_dir = _ops_task_dir(task_id)
    sources_dir = os.path.join(task_dir, "sources")
    outputs_dir = os.path.join(task_dir, "outputs")
    os.makedirs(sources_dir, exist_ok=True)
    os.makedirs(outputs_dir, exist_ok=True)

    if from_task:
        src_main = from_task["astar_main"]
        src_comps = from_task["astar_comps"]
        saved_main = _ops_copy_to_dir(src_main["path"], sources_dir, "main", 0, src_main.get("original_name"))
        saved_comps = [
            _ops_copy_to_dir(item["path"], sources_dir, "comp", i, item.get("original_name"))
            for i, item in enumerate(src_comps)
        ]
        _ops_set_task(task_id, source_task_id=astar_task_id)
    else:
        saved_main = _ops_save_file(main_file, sources_dir, "main", 0)
        saved_comps = [_ops_save_file(f, sources_dir, "comp", i) for i, f in enumerate(comp_files)]

    def _run_output_bg():
        _ops_set_task(task_id, status="running", started_at=_ops_now(), message="Output 生成中")
        try:
            _ops_update_step(task_id, 0, "running", "检查文件")
            _ops_validate_astar_input_columns(saved_main["path"])
            for item in saved_comps:
                _ops_validate_astar_input_columns(item["path"])
            _ops_update_step(task_id, 0, "done", "完成")

            analysis_base = 1

            def _analysis_cb(event, idx=0, detail=""):
                if event == "source_start":
                    _ops_update_step(task_id, analysis_base + idx, "running", detail)
                elif event == "source_done":
                    _ops_update_step(task_id, analysis_base + idx, "done", "完成")
                elif event == "query_start":
                    _ops_update_step(task_id, len(_ops_get_task(task_id).get("steps", [])) - 1, "running", detail)
                elif event == "query_progress":
                    _ops_update_step(task_id, len(_ops_get_task(task_id).get("steps", [])) - 1, "running", detail)

            out_name = f"ops_{task_id}"
            out_path = main_030822.run_analysis(
                saved_main["path"],
                [item["path"] for item in saved_comps],
                output_name=out_name,
                output_dir=outputs_dir,
                progress_cb=_analysis_cb,
                match_config=None,
                post_match_template=rule_template.get("config"),
            )
            final_name = f"output_{rule_template.get('name') or '规则模板'}_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
            final_path = os.path.join(outputs_dir, final_name)
            if os.path.abspath(out_path) != os.path.abspath(final_path):
                shutil.move(out_path, final_path)
            _ops_update_step(task_id, len(_ops_get_task(task_id).get("steps", [])) - 1, "done", "分析完成")
            _ops_set_task(
                task_id,
                status="done",
                ended_at=_ops_now(),
                message="Output 生成完成",
                result_path=final_path,
                result_kind="output_xlsx",
                download_name=final_name,
            )
        except BaseException as e:
            traceback.print_exc()
            running_idx = next((i for i, s in enumerate((_ops_get_task(task_id) or {}).get("steps", [])) if s["status"] == "running"), 0)
            _ops_update_step(task_id, running_idx, "failed", str(e))
            _ops_fail_task(task_id, e)

    threading.Thread(target=_run_output_bg, daemon=True).start()
    return jsonify({"status": "ok", "task_id": task_id})


@app.route('/api/ops/tasks/<task_id>/progress')
def api_ops_task_progress(task_id):
    license_err = _ops_license_error_response()
    if license_err:
        return license_err
    task = _ops_get_task(task_id)
    if not task:
        return jsonify({"status": "error", "message": "任务不存在"}), 404
    return jsonify({"status": "ok", "task": _ops_public_task(task)})


@app.route('/api/ops/tasks/<task_id>/download')
def api_ops_task_download(task_id):
    license_err = _ops_license_error_response()
    if license_err:
        return license_err
    task = _ops_get_task(task_id)
    if not task:
        return jsonify({"status": "error", "message": "任务不存在"}), 404
    path = task.get("result_path") or ""
    if task.get("status") != "done" or not path or not os.path.exists(path):
        return jsonify({"status": "error", "message": "结果文件尚未生成"}), 400
    resp = send_file(path, as_attachment=True, download_name=task.get("download_name") or os.path.basename(path))
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp


@app.route('/api/ops/license-key-status')
def api_ops_license_key_status():
    license_err = _ops_license_error_response()
    if license_err:
        return license_err
    status = _ops_default_private_key_status()
    return jsonify({"status": "ok", **status})


@app.route('/api/ops/license-generate', methods=['POST'])
def api_ops_license_generate():
    license_err = _ops_license_error_response()
    if license_err:
        return license_err
    raw_hwids = (request.form.get("hwids") or "").strip()
    hwids = [x.strip().upper() for x in raw_hwids.replace(",", "\n").replace("，", "\n").splitlines() if x.strip()]
    if not hwids:
        return jsonify({"status": "error", "message": "请填写至少一个 HWID"}), 400
    try:
        days = int((request.form.get("days") or "30").strip())
    except ValueError:
        return jsonify({"status": "error", "message": "有效天数必须是数字"}), 400
    if days <= 0 or days > 3650:
        return jsonify({"status": "error", "message": "有效天数需在 1 到 3650 之间"}), 400

    task = _ops_create_task("license", ["生成 license.dat"], "授权文件生成中")
    task_id = task["task_id"]
    task_dir = _ops_task_dir(task_id)
    out_path = os.path.join(task_dir, "license.dat")
    uploaded_key = request.files.get("private_key")
    try:
        _ops_set_task(task_id, status="running", started_at=_ops_now())
        _ops_update_step(task_id, 0, "running", "签名授权")
        expires = _ops_create_license_file(hwids, days, out_path, uploaded_key=uploaded_key)
        _ops_update_step(task_id, 0, "done", f"到期日 {expires}")
        _ops_set_task(
            task_id,
            status="done",
            ended_at=_ops_now(),
            message=f"license.dat 已生成，到期日 {expires}",
            result_path=out_path,
            result_kind="license_dat",
            download_name="license.dat",
        )
    except BaseException as e:
        traceback.print_exc()
        _ops_update_step(task_id, 0, "failed", str(e))
        _ops_fail_task(task_id, e)
    return jsonify({"status": "ok", "task_id": task_id})


@app.route('/api/ops/package-build', methods=['POST'])
def api_ops_package_build():
    license_err = _ops_license_error_response()
    if license_err:
        return license_err
    target = (request.form.get("target") or "").strip().lower()
    if target not in ("macos", "windows"):
        return jsonify({"status": "error", "message": "请选择 macOS 或 Windows"}), 400

    task = _ops_create_task("package", ["检查环境", "PyArmor 混淆", "PyInstaller 打包", "压缩产物"], f"{target} 打包排队中")
    task_id = task["task_id"]

    def _run_package_bg():
        _ops_set_task(task_id, status="running", started_at=_ops_now(), message=f"{target} 打包中")
        try:
            system = platform.system()
            _ops_update_step(task_id, 0, "running", f"当前系统 {system}")
            if target == "macos":
                if system != "Darwin":
                    raise RuntimeError("macOS .app 需要在 macOS 打包机上执行")
                patch_script = os.path.join(resource_root, "tools", "patch_pyinstaller_site_packages.py")
                if os.path.isfile(patch_script):
                    _ops_run_command(task_id, 0, [sys.executable, patch_script], resource_root)
                else:
                    _ops_update_step(task_id, 0, "done", "未找到 patch 脚本，跳过")
                spec = "ProImage_macOS.spec"
                artifact = os.path.join(resource_root, "dist", "ProImage_AI.app")
                zip_name = f"ProImage_AI_macOS_{time.strftime('%Y%m%d_%H%M%S')}.zip"
            else:
                if system != "Windows":
                    raise RuntimeError("Windows 程序需要在 Windows 打包机上执行")
                _ops_update_step(task_id, 0, "done", "Windows 环境")
                spec = "ProImage_Windows.spec"
                artifact = os.path.join(resource_root, "dist", "ProImage")
                zip_name = f"ProImage_Windows_{time.strftime('%Y%m%d_%H%M%S')}.zip"

            obf_dir = os.path.join(resource_root, "dist", "obfuscated")
            if os.path.isdir(obf_dir):
                shutil.rmtree(obf_dir)
            pyarmor_cmd = _ops_pyarmor_command() + ["gen", "-O", os.path.join("dist", "obfuscated")] + _OPS_PYARMOR_FILES
            try:
                _ops_run_command(task_id, 1, pyarmor_cmd, resource_root)
                _ops_verify_pyarmor_output(obf_dir)
                _ops_update_step(task_id, 1, "done", f"完成，已生成 {len(_OPS_PYARMOR_FILES)} 个混淆文件")
            except BaseException as e:
                if os.path.isdir(obf_dir):
                    shutil.rmtree(obf_dir)
                detail = str(e).splitlines()[-1] if str(e).splitlines() else str(e)
                _ops_update_step(task_id, 1, "done", f"混淆失败，已清理并改用源码模式：{detail[:180]}")

            _ops_run_command(task_id, 2, [sys.executable, "-m", "PyInstaller", "-y", spec], resource_root)
            if not os.path.exists(artifact):
                raise RuntimeError(f"打包产物不存在：{artifact}")
            if target == "macos":
                _ops_run_command(task_id, 2, ["xattr", "-cr", artifact], resource_root)
                _ops_run_command(task_id, 2, ["codesign", "--force", "--deep", "--sign", "-", artifact], resource_root)
            task_dir = _ops_task_dir(task_id)
            zip_path = os.path.join(task_dir, zip_name)
            _ops_update_step(task_id, 3, "running", "压缩产物")
            _ops_zip_path(artifact, zip_path)
            _ops_update_step(task_id, 3, "done", "完成")
            _ops_set_task(
                task_id,
                status="done",
                ended_at=_ops_now(),
                message="打包完成",
                result_path=zip_path,
                result_kind="package_zip",
                download_name=zip_name,
            )
        except BaseException as e:
            traceback.print_exc()
            running_idx = next((i for i, s in enumerate((_ops_get_task(task_id) or {}).get("steps", [])) if s["status"] == "running"), 0)
            _ops_update_step(task_id, running_idx, "failed", str(e))
            _ops_fail_task(task_id, e)

    threading.Thread(target=_run_package_bg, daemon=True).start()
    return jsonify({"status": "ok", "task_id": task_id})

@app.route('/api/projects', methods=['GET', 'POST'])
def handle_projects():
    if request.method == 'POST':
        name = request.form.get('name')
        if not name: return jsonify({"status": "error", "message": "Project name is required"}), 400

        match_config_json = (request.form.get('match_config_json') or "").strip()
        raw_rt = (request.form.get("rule_template_id") or "").strip()
        try:
            rule_template_id = int(raw_rt) if raw_rt else None
        except ValueError:
            rule_template_id = None

        main_file = request.files.get('main_file')
        comp_files = request.files.getlist('comp_files')
        
        if not main_file or not main_file.filename:
            return jsonify({"status": "error", "message": "Main store file is required"}), 400
        
        valid_comp_files = [f for f in comp_files if f.filename]
        if not valid_comp_files:
            return jsonify({"status": "error", "message": "At least one competitor store file is required"}), 400

        err = _validate_upload(main_file, "主店文件")
        if err: return jsonify({"status": "error", "message": err}), 400
        for f in valid_comp_files:
            err = _validate_upload(f, f"竞店文件 ({f.filename})")
            if err: return jsonify({"status": "error", "message": err}), 400
        result_file = request.files.get('result_file')
        if result_file and result_file.filename:
            err = _validate_upload(result_file, "结果文件")
            if err: return jsonify({"status": "error", "message": err}), 400

        # Temporary PID for directory naming
        temp_pid = int(time.time())
        proj_dir = os.path.join(data_root, "uploads", f"project_{temp_pid}")
        sources_dir = os.path.join(proj_dir, "sources")
        os.makedirs(sources_dir, exist_ok=True)
        
        # Save files with role prefixes. Main and competitor uploads can have the
        # same original filename; storing them directly in one directory would
        # let the later competitor save overwrite the main file.
        main_saved_name = "main__" + _safe_upload_filename(main_file.filename, "main.xlsx")
        main_path = os.path.join(sources_dir, main_saved_name)
        main_file.save(main_path)
        main_store_name = main_file.filename.replace(".xlsx", "").replace(".xls", "")
        
        # Save Competitor Store Files
        comp_infos = []
        comp_paths = []
        for idx, f in enumerate(valid_comp_files):
            comp_saved_name = f"comp_{idx}__" + _safe_upload_filename(f.filename, f"comp_{idx}.xlsx")
            path = os.path.join(sources_dir, comp_saved_name)
            f.save(path)
            comp_paths.append(path)
            comp_infos.append({"path": path, "store_name": f.filename.replace(".xlsx", "").replace(".xls", "")})
        
        manual_result_path = None
        if result_file and result_file.filename:
            outputs_dir = os.path.join(proj_dir, "outputs")
            os.makedirs(outputs_dir, exist_ok=True)
            manual_result_path = os.path.join(outputs_dir, result_file.filename)
            result_file.save(manual_result_path)

        is_manual = bool(manual_result_path)
        pid = dm.create_project(
            name,
            {"path": main_path, "store_name": main_store_name},
            comp_infos,
            status='ready' if is_manual else 'analyzing',
            match_config_json=match_config_json,
            rule_template_id=rule_template_id,
        )

        # Rename temp directory to real PID
        real_proj_dir = os.path.join(data_root, "uploads", f"project_{pid}")
        if os.path.exists(real_proj_dir): shutil.rmtree(real_proj_dir)
        os.rename(proj_dir, real_proj_dir)

        with dm._db_lock:
            with dm._get_conn() as conn:
                conn.execute("UPDATE project_files SET local_path = REPLACE(local_path, ?, ?) WHERE project_id = ?",
                            (f"project_{temp_pid}", f"project_{pid}", pid))

        dirs = dm._ensure_project_dirs(pid)
        final_main_path = main_path.replace(f"project_{temp_pid}", f"project_{pid}")
        final_comp_paths = [p.replace(f"project_{temp_pid}", f"project_{pid}") for p in comp_paths]
        final_manual_result_path = manual_result_path.replace(f"project_{temp_pid}", f"project_{pid}") if manual_result_path else None

        if is_manual:
            # Sync path: copy result → activate → import → redirect to dashboard
            output_file = os.path.join(dirs["outputs"], f"output_{pid}.xlsx")
            shutil.copy(final_manual_result_path, output_file)
            dm.activate_project(pid)
            return jsonify({"status": "success", "project_id": pid})

        # Async path: return immediately, run analysis in background thread
        use_ai = request.form.get('use_ai') == 'on'
        api_key = request.form.get('api_key')
        ai_model_name = (request.form.get('ai_model_name') or "").strip()
        kimi_api_key = (request.form.get('kimi_api_key') or "").strip()
        kimi_model_name = (request.form.get('kimi_model_name') or "").strip()

        main_name = os.path.basename(final_main_path).replace('.xlsx','').replace('.xls','')
        comp_names = [os.path.basename(p).replace('.xlsx','').replace('.xls','') for p in final_comp_paths]

        def _run_analysis_bg():
            _t0 = time.time()
            print(f"[BG] Project {pid} thread started at {time.strftime('%H:%M:%S')}", flush=True)
            has_ai = bool(use_ai and api_key)
            prog = _init_progress(pid, has_ai, main_name, comp_names)
            ai_file_count = (1 + len(comp_names)) if has_ai else 0
            try:
                if has_ai:
                    all_ai_paths = [final_main_path] + final_comp_paths
                    try:
                        _ai_gap = int(os.environ.get("PROIMAGE_AI_INTER_FILE_SLEEP_SEC", "8") or "8")
                    except ValueError:
                        _ai_gap = 8
                    _ai_gap = max(0, _ai_gap)
                    for fi, fp in enumerate(all_ai_paths):
                        _update_step(pid, fi, "running")
                        def _ai_cb(batch, total, _fi=fi):
                            _update_step(pid, _fi, "running", f"batch {batch}/{total}")
                        extract_info_ai2.process_file_ai(
                            fp,
                            api_key,
                            progress_cb=_ai_cb,
                            model_name=ai_model_name,
                            fallback_api_key=kimi_api_key or None,
                            fallback_model=kimi_model_name or None,
                        )
                        _update_step(pid, fi, "done")
                        if fi + 1 < len(all_ai_paths) and _ai_gap > 0:
                            print(
                                f"[BG] Project {pid} AI: 间歇 {_ai_gap}s 后处理下一文件（减轻 Gemini 连续请求 429/限流）",
                                flush=True,
                            )
                            time.sleep(_ai_gap)
                    print(f"[BG] Project {pid} AI extraction done in {time.time()-_t0:.1f}s", flush=True)

                analysis_base = ai_file_count
                def _analysis_cb(event, idx=0, detail=""):
                    if event == "source_start":
                        _update_step(pid, analysis_base + idx, "running", detail)
                    elif event == "source_done":
                        _update_step(pid, analysis_base + idx, "done")
                    elif event == "query_start":
                        _update_step(pid, len(prog["steps"]) - 1, "running", detail)
                    elif event == "query_progress":
                        _update_step(pid, len(prog["steps"]) - 1, "running", detail)

                print(f"[BG] Project {pid} starting run_analysis...", flush=True)
                _pmt = dm.get_post_match_template_for_project(pid)
                main_030822.run_analysis(
                    final_main_path, final_comp_paths,
                    output_name=str(pid), output_dir=dirs["outputs"],
                    progress_cb=_analysis_cb,
                    match_config=match_config_json,
                    post_match_template=_pmt,
                )
                _update_step(pid, len(prog["steps"]) - 1, "done", "分析完成")
                dm.update_project_status(pid, 'ready')
                print(f"[BG] Project {pid} analysis complete in {time.time()-_t0:.1f}s", flush=True)
            except BaseException as e:
                traceback.print_exc()
                try:
                    dm.update_project_status(pid, 'failed')
                except Exception:
                    pass
                print(f"[BG] Project {pid} FAILED ({type(e).__name__}: {e}) after {time.time()-_t0:.1f}s", flush=True)
            finally:
                _schedule_clear_progress(pid)

        threading.Thread(target=_run_analysis_bg, daemon=True).start()
        return jsonify({"status": "success", "project_id": pid})
        
    return jsonify(dm.list_projects())

@app.route('/api/projects/<int:pid>/activate', methods=['POST'])
def activate_project(pid):
    projects = dm.list_projects()
    proj = next((p for p in projects if p['id'] == pid), None)
    if not proj:
        return jsonify({"status": "error", "message": "项目不存在"}), 404
    if proj.get('status') == 'analyzing':
        return jsonify({"status": "error", "message": "该项目正在分析中，请等待完成"}), 400
    if proj.get('status') == 'failed':
        return jsonify({"status": "error", "message": "该项目分析失败，请删除后重新创建"}), 400
    dm.activate_project(pid)
    return jsonify({"status": "success"})

@app.route('/api/projects/<int:pid>', methods=['DELETE'])
def delete_project(pid):
    dm.delete_project(pid)
    return jsonify({"status": "success"})

@app.route("/match-rules")
def match_rules_page():
    is_valid, _ = check_license()
    if not is_valid:
        return render_template("activate.html", hwid=CURRENT_HWID)
    return render_template("match_rules.html")

@app.route("/match-rules/new", methods=["GET"])
def match_rule_new_page():
    is_valid, _ = check_license()
    if not is_valid:
        return render_template("activate.html", hwid=CURRENT_HWID)
    return render_template("match_rule_edit.html", template_id=0, template_name="")

@app.route("/match-rules/<int:tid>", methods=["GET"])
def match_rule_edit_page(tid):
    is_valid, _ = check_license()
    if not is_valid:
        return render_template("activate.html", hwid=CURRENT_HWID)
    t = dm.get_rule_template(tid) if tid else None
    if tid and not t:
        return "规则不存在", 404
    return render_template("match_rule_edit.html", template_id=tid, template_name=(t or {}).get("name", ""))

@app.route("/api/rule-templates", methods=["GET", "POST"])
def api_rule_templates():
    if request.method == "GET":
        return jsonify({"status": "ok", "items": dm.list_rule_templates()})
    d = request.get_json(silent=True) or {}
    name = (d.get("name") or "").strip()
    if not name:
        return jsonify({"status": "error", "message": "模板名称不能为空"}), 400
    desc = (d.get("description") or "").strip()
    config = d.get("config")
    if not isinstance(config, dict):
        return jsonify({"status": "error", "message": "config 须为对象"}), 400
    rid = dm.create_rule_template(name, desc, config)
    return jsonify({"status": "ok", "id": rid})

@app.route("/api/rule-templates/<int:tid>", methods=["GET", "PUT", "DELETE"])
def api_rule_template_one(tid):
    if request.method == "GET":
        t = dm.get_rule_template(tid)
        if not t:
            return jsonify({"status": "error", "message": "不存在"}), 404
        return jsonify({"status": "ok", "item": t})
    if request.method == "DELETE":
        ok, err = dm.delete_rule_template(tid)
        if not ok:
            return jsonify({"status": "error", "message": err or "删除失败"}), 400
        return jsonify({"status": "ok"})
    d = request.get_json(silent=True) or {}
    name = (d.get("name") or "").strip()
    if not name:
        return jsonify({"status": "error", "message": "模板名称不能为空"}), 400
    desc = (d.get("description") or "").strip()
    config = d.get("config")
    if not isinstance(config, dict):
        return jsonify({"status": "error", "message": "config 须为对象"}), 400
    if not dm.update_rule_template(tid, name, desc, config):
        return jsonify({"status": "error", "message": "更新失败"}), 400
    return jsonify({"status": "ok"})


@app.route("/api/rule-category-template")
def api_rule_category_template():
    if os.path.isfile(DEFAULT_RULE_CATEGORIES_XLSX):
        return send_file(
            DEFAULT_RULE_CATEGORIES_XLSX,
            as_attachment=True,
            download_name="类目配置模板.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    wb = Workbook()
    ws = wb.active
    ws.title = "类目模板"
    ws.append(["美团一级类目", "美团二级类目", "美团三级类目"])
    ws.append(["饮料", "碳酸饮料", "可乐"])
    ws.append(["饮料", "碳酸饮料", "雪碧"])
    ws.append(["休闲零食", "膨化食品", "薯片"])
    ws.append(["生鲜水果", "热带水果", "香蕉"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="类目配置模板.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/rule-categories/parse", methods=["POST"])
def api_rule_categories_parse():
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"status": "error", "message": "请上传类目文件"}), 400
    err = _validate_upload(f, "类目文件")
    if err:
        return jsonify({"status": "error", "message": err}), 400
    try:
        rows = _excel_file_to_rows(f)
        tree = _build_category_tree(rows)
        if not tree["items"]:
            return jsonify({"status": "error", "message": "未读取到有效的三级类目，请检查模板列名与内容"}), 400
        return jsonify({"status": "ok", "tree": tree})
    except Exception as e:
        return jsonify({"status": "error", "message": f"解析失败：{e}"}), 400


@app.route("/api/rule-categories/default", methods=["GET"])
def api_rule_categories_default():
    """规则编辑页首次进入时使用的默认一级/二/三级类目树（与 data/default_meituan_categories.xlsx 一致）。"""
    if not os.path.isfile(DEFAULT_RULE_CATEGORIES_XLSX):
        return jsonify({"status": "error", "message": "未配置默认类目文件"}), 404
    try:
        rows = _excel_path_to_rows(DEFAULT_RULE_CATEGORIES_XLSX)
        tree = _build_category_tree(rows)
        if not tree["items"]:
            return jsonify({"status": "error", "message": "默认类目文件无有效三级类目"}), 500
        return jsonify({"status": "ok", "tree": tree})
    except Exception as e:
        return jsonify({"status": "error", "message": f"解析失败：{e}"}), 500


@app.route("/api/rule-categories/bucket-tags", methods=["GET"])
def api_rule_categories_bucket_tags():
    """规则编辑页：一级类目快捷标签（枚举），用于批量勾选其下三级类目。"""
    if not os.path.isfile(CATEGORY_L1_BUCKET_TAGS_JSON):
        return jsonify({"status": "ok", "version": 0, "tags": []})
    try:
        with open(CATEGORY_L1_BUCKET_TAGS_JSON, "r", encoding="utf-8") as f:
            data = json.load(f)
        tags = data.get("tags") if isinstance(data, dict) else []
        if not isinstance(tags, list):
            tags = []
        return jsonify({"status": "ok", "version": int(data.get("version", 1) if isinstance(data, dict) else 1), "tags": tags})
    except Exception as e:
        return jsonify({"status": "error", "message": f"读取标签配置失败：{e}"}), 500


@app.route("/api/config", methods=['GET'])
def get_config():
    return jsonify({
        "main_store": dm.main_store_name, "target_file": dm.target_file, "output_file": dm.output_file,
        "source_files": dm.source_files, "stores": [{"id": str(i), "name": n, "path": dm.source_files[i]} for i, n in enumerate(dm.store_names)]
    })

@app.route('/api/grid_data')
def get_grid_data():
    page = request.args.get('page', 1, type=int)
    limit = request.args.get('limit', 50, type=int)
    search = request.args.get('search', "")
    mode = request.args.get('mode', "all")
    filters_json = request.args.get('filters', "{}")
    sort_field = request.args.get('sort_field', "")
    sort_order = request.args.get('sort_order', "desc")
    negative_sales = request.args.get('negative_sales', "0") == "1"
    return jsonify(dm.get_paginated_grid(
        page=page, limit=limit, search=search, mode=mode,
        filters_json=filters_json, sort_field=sort_field, sort_order=sort_order,
        negative_sales_only=negative_sales,
    ))

@app.route('/api/store_products/<store_id>')
def get_store_products(store_id):
    return jsonify(dm.get_store_products(store_id))


@app.route('/api/unlinked_items')
def get_unlinked_items():
    page = request.args.get('page', 1, type=int)
    limit = request.args.get('limit', 30, type=int)
    search = request.args.get('search', "")
    category3 = request.args.get('category3', "")
    sort_store_id = request.args.get('sort_store_id', "")
    sort_order = request.args.get('sort_order', "desc")
    filters_json = request.args.get('filters', "{}")
    negative_sales = request.args.get('negative_sales', "0") == "1"
    return jsonify(dm.get_unlinked_pool_page(
        page=page, limit=limit, search=search, category3=category3,
        sort_store_id=sort_store_id, sort_order=sort_order,
        filters_json=filters_json, negative_sales_only=negative_sales,
    ))

@app.route('/api/main_products')
def get_main_products():
    page = request.args.get('page', 1, type=int)
    limit = request.args.get('limit', 50, type=int)
    search = request.args.get('search', "")
    return jsonify(dm.get_main_products_page(page=page, limit=limit, search=search))

@app.route('/api/eliminate', methods=['POST'])
def eliminate():
    d = request.json
    main_sku_id = d.get('main_sku_id')
    if not main_sku_id:
        return jsonify({"status": "error", "message": "Missing main_sku_id"}), 400
    dm.eliminate_product(main_sku_id, d.get('status', 1))
    return jsonify({"status": "success"})

@app.route('/api/toggle_handled', methods=['POST'])
def toggle_handled():
    d = request.json
    sku_id = d.get('main_sku_id')
    if not sku_id:
        return jsonify({"status": "error", "message": "Missing main_sku_id"}), 400
    dm.toggle_handled(sku_id, d.get('handled', True))
    return jsonify({"status": "success"})

@app.route('/api/toggle_ref', methods=['POST'])
def toggle_ref():
    d = request.json
    sku_id = d.get('main_sku_id')
    field = d.get('field')  # 'name' or 'image'
    store_id = d.get('store_id', '')
    if not sku_id or field not in ('name', 'image'):
        return jsonify({"status": "error", "message": "Missing params"}), 400
    dm.set_ref(sku_id, field, store_id)
    return jsonify({"status": "success"})

@app.route('/api/toggle_add', methods=['POST'])
def toggle_add():
    d = request.json
    store_id = d.get('store_id')
    comp_sku_id = d.get('sku_id')
    if store_id is None or not comp_sku_id:
        return jsonify({"status": "error", "message": "Missing store_id or sku_id"}), 400
    ok = dm.mark_as_new(store_id, comp_sku_id, d.get('is_new', True))
    if not ok:
        return jsonify({"status": "error", "message": "未找到可标记的商品"}), 400
    return jsonify({"status": "success"})

@app.route('/api/price_match', methods=['POST'])
def price_match():
    d = request.json
    main_sku_id = d.get('main_sku_id')
    store_id = d.get('store_id')
    if not main_sku_id or store_id is None:
        return jsonify({"status": "error", "message": "Missing params"}), 400
    result = dm.price_match(main_sku_id, store_id)
    if not result:
        return jsonify({"status": "error", "message": "未找到可跟价的商品"}), 400
    return jsonify({"status": "success", **result})

@app.route('/api/clear_price_match', methods=['POST'])
def clear_price_match():
    d = request.json
    main_sku_id = d.get('main_sku_id')
    if not main_sku_id:
        return jsonify({"status": "error", "message": "Missing main_sku_id"}), 400
    dm.clear_price_match(main_sku_id)
    return jsonify({"status": "success"})

@app.route('/api/manual_link', methods=['POST'])
def manual_link():
    d = request.json
    main_sku_id = d.get('main_sku_id')
    store_id = d.get('store_id')
    comp_sku_id = d.get('comp_sku_id')
    if not main_sku_id or store_id is None or not comp_sku_id:
        return jsonify({"status": "error", "message": "Missing params"}), 400
    dm.manual_link(main_sku_id, store_id, comp_sku_id)
    return jsonify({"status": "success"})

@app.route('/api/unlink', methods=['POST'])
def unlink():
    d = request.json
    main_sku_id = d.get('main_sku_id')
    store_id = d.get('store_id')
    if not main_sku_id or store_id is None:
        return jsonify({"status": "error", "message": "Missing params"}), 400
    dm.unlink_product(main_sku_id, store_id)
    return jsonify({"status": "success"})

@app.route('/api/update_cell', methods=['POST'])
def update_cell():
    d = request.json
    main_sku_id = d.get('main_sku_id')
    if not main_sku_id:
        return jsonify({"status": "error", "message": "Missing main_sku_id"}), 400
    dm.update_cell(main_sku_id, {d.get('column'): d.get('value')})
    return jsonify({"status": "success"})

@app.route('/img/<path:filename>')
def serve_img(filename):
    return send_from_directory(os.path.join(data_root, "img"), filename)


@app.route('/api/export')
def export_data():
    p = dm.save_separate_exports()
    resp = send_file(p, as_attachment=True)
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"; resp.headers["Pragma"] = "no-cache"; resp.headers["Expires"] = "0"
    return resp

@app.route('/api/export_new')
def export_new_data():
    p = dm.export_new_items()
    resp = send_file(p, as_attachment=True)
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"; resp.headers["Pragma"] = "no-cache"; resp.headers["Expires"] = "0"
    return resp

@app.route('/api/projects/<int:pid>/progress')
def get_analysis_progress(pid):
    with _progress_lock:
        prog = _analysis_progress.get(pid)
    if not prog:
        return jsonify({"available": False})
    steps = prog["steps"]
    elapsed = time.time() - prog["started_at"]
    done_count = sum(1 for s in steps if s["status"] == "done")
    total = len(steps)
    pct = int(done_count / total * 100) if total else 0
    running_idx = next((i for i, s in enumerate(steps) if s["status"] == "running"), -1)
    if running_idx >= 0:
        pct = int((done_count + 0.5) / total * 100)
    done_durations = [s["ended_at"] - s["started_at"] for s in steps
                      if s.get("started_at") and s.get("ended_at")]
    avg_step = (sum(done_durations) / len(done_durations)) if done_durations else 0
    remaining_steps = total - done_count - (1 if running_idx >= 0 else 0)
    running_elapsed = (time.time() - steps[running_idx]["started_at"]) if running_idx >= 0 and steps[running_idx].get("started_at") else 0
    est_remaining = max(0, avg_step - running_elapsed) + remaining_steps * avg_step if avg_step > 0 else 0
    out_steps = []
    for s in steps:
        item = {"label": s["label"], "status": s["status"], "detail": s.get("detail", "")}
        if s.get("started_at") and s.get("ended_at"):
            item["duration_s"] = round(s["ended_at"] - s["started_at"], 1)
        elif s.get("started_at"):
            item["running_s"] = round(time.time() - s["started_at"], 1)
        out_steps.append(item)
    return jsonify({
        "available": True, "elapsed_s": round(elapsed, 1),
        "pct": pct, "estimated_remaining_s": round(est_remaining, 1),
        "done_count": done_count, "total_steps": total,
        "steps": out_steps,
    })

@app.route('/api/debug/threads')
def debug_threads():
    import io
    buf = io.StringIO()
    buf.write(f"Time: {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
    buf.write(f"Active threads: {threading.active_count()}\n\n")
    frames = sys._current_frames()
    for tid, frame in frames.items():
        tname = "unknown"
        for t in threading.enumerate():
            if t.ident == tid:
                tname = t.name
                break
        buf.write(f"--- Thread {tid} ({tname}) ---\n")
        traceback.print_stack(frame, file=buf)
        buf.write("\n")
    return buf.getvalue(), 200, {'Content-Type': 'text/plain; charset=utf-8'}

if __name__ == '__main__':
    if getattr(sys, 'frozen', False):
        import webbrowser
        threading.Timer(1.5, lambda: webbrowser.open('http://127.0.0.1:5001')).start()
    app.run(debug=False, port=5001)
