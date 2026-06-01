import os
import sys
import shutil
import time
import threading
import traceback
import io
import json
import uuid
import zipfile
import subprocess
import platform
import datetime
from flask import Blueprint, request, jsonify, render_template, send_file
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from field_registry import detect_field_mapping

# These will be initialized by app.py
dm = None
resource_root = None
data_root = None
check_license = None
CURRENT_HWID = None
extract_info_ai2 = None
main_030822 = None
_validate_upload = None
_safe_upload_filename = None

ops_bp = Blueprint('ops', __name__)

_ops_tasks = {}
_ops_lock = threading.Lock()

def init_ops(app_obj, dm_obj, res_root, dat_root, license_fn, hwid, ai_mod, main_mod, validate_fn, safe_name_fn):
    global dm, resource_root, data_root, check_license, CURRENT_HWID, extract_info_ai2, main_030822
    global _validate_upload, _safe_upload_filename
    dm = dm_obj
    resource_root = res_root
    data_root = dat_root
    check_license = license_fn
    CURRENT_HWID = hwid
    extract_info_ai2 = ai_mod
    main_030822 = main_mod
    _validate_upload = validate_fn
    _safe_upload_filename = safe_name_fn
    
    import app_ops_extra
    ctx = {
        "dm": dm, "resource_root": resource_root, "data_root": data_root, 
        "check_license": check_license, "CURRENT_HWID": CURRENT_HWID,
        "license_error_response": _ops_license_error_response,
        "default_private_key_status": _ops_default_private_key_status,
        "create_task": _ops_create_task, "get_task": _ops_get_task, "set_task": _ops_set_task,
        "update_step": _ops_update_step, "fail_task": _ops_fail_task, "now": _ops_now,
        "task_dir": _ops_task_dir, "create_license_file": _ops_create_license_file,
        "run_command": _ops_run_command,
        "zip_path": _ops_zip_path,
        "extract_info_ai2": extract_info_ai2, "main_030822": main_030822,
        "dm": dm,
        "collect_astar_source_files": _ops_collect_astar_source_files,
        "validate_source_uploads": _ops_validate_source_uploads,
        "file_label": _ops_file_label, "save_file": _ops_save_file,
        "copy_to_dir": _ops_copy_to_dir, "validate_astar_input_columns": _ops_validate_astar_input_columns,
        "zip_files": _ops_zip_files, "rule_template_from_request": _ops_rule_template_from_request,
        "public_astar_file_choices": _ops_public_astar_file_choices,
        "get_astar_choice": _ops_get_astar_choice, "validate_excel_uploads": _ops_validate_excel_uploads,
        "validate_upload": _validate_upload,
    }
    app_ops_extra.init_ops_extra(ctx)
    app_obj.register_blueprint(app_ops_extra.extra_bp)
    import app_ops_tasks
    app_ops_tasks.init_ops_tasks(ctx)
    app_obj.register_blueprint(app_ops_tasks.tasks_bp)

def _ops_now():
    return time.strftime("%Y-%m-%d %H:%M:%S")

def _ops_task_dir(task_id):
    return os.path.join(data_root, "uploads", "ops_tasks", task_id)

def _ops_public_task(task):
    astar_files = []
    for item in _ops_public_astar_file_choices(task):
        astar_files.append({
            "index": item.get("index"),
            "original_name": item.get("original_name", ""),
        })
    return {
        "task_id": task.get("task_id"),
        "kind": task.get("kind"),
        "status": task.get("status"),
        "message": task.get("message", ""),
        "error": task.get("error", ""),
        "created_at": task.get("created_at"),
        "started_at": task.get("started_at"),
        "ended_at": task.get("ended_at"),
        "steps": task.get("steps", []),
        "download_ready": bool(task.get("result_path") and os.path.exists(task.get("result_path", ""))),
        "download_name": task.get("download_name", ""),
        "result_kind": task.get("result_kind", ""),
        "source_task_id": task.get("source_task_id", ""),
        "quality_summary": task.get("quality_summary", {}),
        "astar_files": astar_files,
    }

def _ops_create_task(kind, steps, message=""):
    task_id = uuid.uuid4().hex[:12]
    task = {
        "task_id": task_id,
        "kind": kind,
        "status": "pending",
        "message": message,
        "error": "",
        "created_at": _ops_now(),
        "started_at": None,
        "ended_at": None,
        "steps": [{"label": s, "status": "pending", "detail": ""} for s in steps],
        "result_path": "",
        "download_name": "",
        "result_kind": "",
    }
    with _ops_lock:
        _ops_tasks[task_id] = task
    return task

def _ops_get_task(task_id):
    with _ops_lock:
        return _ops_tasks.get(task_id)

def _ops_set_task(task_id, **kwargs):
    with _ops_lock:
        task = _ops_tasks.get(task_id)
        if not task:
            return
        task.update(kwargs)

def _ops_update_step(task_id, idx, status, detail=""):
    with _ops_lock:
        task = _ops_tasks.get(task_id)
        if not task or idx < 0 or idx >= len(task.get("steps", [])):
            return
        step = task["steps"][idx]
        step["status"] = status
        step["detail"] = detail or ""
        now = time.time()
        if status == "running" and not step.get("started_at"):
            step["started_at"] = now
        if status in ("done", "failed") and not step.get("ended_at"):
            step["ended_at"] = now

def _ops_fail_task(task_id, err):
    _ops_set_task(task_id, status="failed", error=str(err), ended_at=_ops_now())

def _ops_file_label(file_storage, fallback):
    name = (getattr(file_storage, "filename", "") or "").strip()
    return name or fallback

def _ops_safe_filename(filename, fallback):
    ext = os.path.splitext(filename or "")[1].lower() or ".xlsx"
    safe = _safe_upload_filename(filename, fallback)
    if not os.path.splitext(safe)[1]:
        safe += ext
    return safe

def _ops_save_file(file_storage, dest_dir, prefix, idx=0):
    os.makedirs(dest_dir, exist_ok=True)
    original = _ops_file_label(file_storage, f"{prefix}_{idx}.xlsx")
    safe = _ops_safe_filename(original, f"{prefix}_{idx}.xlsx")
    base, ext = os.path.splitext(safe)
    filename = f"{prefix}_{idx}__{base}{ext}"
    path = os.path.join(dest_dir, filename)
    n = 1
    while os.path.exists(path):
        filename = f"{prefix}_{idx}__{base}_{n}{ext}"
        path = os.path.join(dest_dir, filename)
        n += 1
    file_storage.save(path)
    return {"path": path, "original_name": original, "safe_name": filename}

def _ops_validate_excel_uploads(main_file, comp_files):
    if not main_file or not main_file.filename:
        return "请上传主店文件"
    err = _validate_upload(main_file, "主店文件")
    if err:
        return err
    valid_comp_files = [f for f in comp_files if f and f.filename]
    if not valid_comp_files:
        return "请至少上传一个竞店文件"
    for f in valid_comp_files:
        err = _validate_upload(f, f"竞店文件 ({f.filename})")
        if err:
            return err
    return None

def _ops_validate_source_uploads(source_files):
    files = [f for f in source_files if f and f.filename]
    if not files:
        return "请至少上传一个原始文件"
    for f in files:
        err = _validate_upload(f, f"原始文件 ({f.filename})")
        if err:
            return err
    return None

def _ops_collect_astar_source_files():
    source_files = [f for f in request.files.getlist('source_files') if f and f.filename]
    if source_files:
        return source_files
    main_file = request.files.get('main_file')
    comp_files = [f for f in request.files.getlist('comp_files') if f and f.filename]
    return ([main_file] if main_file and main_file.filename else []) + comp_files

def _ops_public_astar_file_choices(task):
    files = task.get("astar_files") or []
    if files:
        return files
    legacy = []
    if task.get("astar_main"):
        legacy.append(task["astar_main"])
    legacy.extend(task.get("astar_comps") or [])
    for idx, item in enumerate(legacy):
        item.setdefault("index", idx)
    return legacy

def _ops_get_astar_choice(task, raw_idx):
    try:
        idx = int(str(raw_idx).strip())
    except (TypeError, ValueError):
        raise ValueError("请选择有效的 A* 文件")
    for item in _ops_public_astar_file_choices(task):
        if int(item.get("index", -1)) == idx:
            return item
    raise ValueError("选择的 A* 文件不存在")

def _ops_validate_astar_input_columns(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        headers = {str(c).strip() for c in (row or []) if c is not None}
        mapping = detect_field_mapping(headers, standards=["商品名称", "规格名称"])
        if not mapping.get("商品名称", {}).get("column") or not mapping.get("规格名称", {}).get("column"):
            raise ValueError("缺少必需列：商品名称 + 规格/规格名称")
    finally:
        wb.close()

def _ops_copy_to_dir(src, dest_dir, prefix, idx, original_name=None):
    os.makedirs(dest_dir, exist_ok=True)
    original = original_name or os.path.basename(src)
    safe = _ops_safe_filename(original, f"{prefix}_{idx}.xlsx")
    base, ext = os.path.splitext(safe)
    filename = f"{prefix}_{idx}__{base}{ext}"
    dest = os.path.join(dest_dir, filename)
    n = 1
    while os.path.exists(dest):
        filename = f"{prefix}_{idx}__{base}_{n}{ext}"
        dest = os.path.join(dest_dir, filename)
        n += 1
    shutil.copy2(src, dest)
    return {"path": dest, "original_name": original, "safe_name": filename}

def _ops_zip_files(files, zip_path):
    os.makedirs(os.path.dirname(zip_path), exist_ok=True)
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        used = set()
        for item in files:
            arc = item.get("arcname") or item.get("original_name") or os.path.basename(item["path"])
            arc = arc.replace("/", "_").replace("\\", "_")
            if arc in used:
                base, ext = os.path.splitext(arc)
                n = 1
                while f"{base}_{n}{ext}" in used:
                    n += 1
                arc = f"{base}_{n}{ext}"
            used.add(arc)
            zf.write(item["path"], arcname=arc)

def _ops_rule_template_from_request(raw_rule_id):
    templates = dm.list_rule_templates()
    tid = None
    try:
        tid = int(raw_rule_id) if str(raw_rule_id or "").strip() else None
    except ValueError:
        tid = None
    if not tid:
        prod = next((t for t in templates if t.get("name") == "生产规则V1"), None)
        tid = int(prod["id"]) if prod else (int(templates[0]["id"]) if templates else None)
    t = dm.get_rule_template(tid) if tid else None
    if not t:
        raise ValueError("未找到可用的类目规则模板")
    return t

def _ops_load_private_key(uploaded_key=None):
    from cryptography.hazmat.primitives import serialization

    key_bytes = None
    if uploaded_key and uploaded_key.filename:
        key_bytes = uploaded_key.read()
        uploaded_key.seek(0)
    else:
        for p in (
            os.path.join(resource_root, "vendor", "private_key.pem"),
            os.path.join(data_root, "vendor", "private_key.pem"),
            os.path.join(resource_root, "private_key.pem"),
            os.path.join(data_root, "private_key.pem"),
        ):
            if os.path.isfile(p):
                with open(p, "rb") as f:
                    key_bytes = f.read()
                break
    if not key_bytes:
        raise ValueError("未找到 private_key.pem；请上传私钥文件，或放到 vendor/private_key.pem")
    return serialization.load_pem_private_key(key_bytes, password=None)

def _ops_default_private_key_status():
    candidates = (
        (os.path.join(resource_root, "vendor", "private_key.pem"), "vendor/private_key.pem"),
        (os.path.join(data_root, "vendor", "private_key.pem"), "vendor/private_key.pem"),
        (os.path.join(resource_root, "private_key.pem"), "private_key.pem"),
        (os.path.join(data_root, "private_key.pem"), "private_key.pem"),
    )
    for path, label in candidates:
        if os.path.isfile(path):
            return {"configured": True, "path_label": label}
    return {"configured": False, "path_label": ""}

def _ops_create_license_file(hwids, expires_days, out_path, uploaded_key=None):
    from cryptography.hazmat.primitives import hashes
    from cryptography.hazmat.primitives.asymmetric import padding
    import base64

    private_key = _ops_load_private_key(uploaded_key)
    expires = (datetime.datetime.now() + datetime.timedelta(days=int(expires_days))).strftime("%Y-%m-%d")
    data = {"hwids": hwids, "expires": expires, "version": "1.0"}
    data_json = json.dumps(data, ensure_ascii=False)
    data_b64 = base64.b64encode(data_json.encode()).decode()
    signature = private_key.sign(
        data_json.encode(),
        padding.PSS(mgf=padding.MGF1(hashes.SHA256()), salt_length=padding.PSS.MAX_LENGTH),
        hashes.SHA256(),
    )
    sig_b64 = base64.b64encode(signature).decode()
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(f"{data_b64}.{sig_b64}")
    return expires

def _ops_run_command(task_id, step_idx, cmd, cwd):
    _ops_update_step(task_id, step_idx, "running", " ".join(cmd))
    proc = subprocess.Popen(
        cmd,
        cwd=cwd,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        bufsize=1,
    )
    last = ""
    if proc.stdout:
        for line in proc.stdout:
            line = line.rstrip()
            if not line:
                continue
            last = line[-240:]
            _ops_update_step(task_id, step_idx, "running", last)
    code = proc.wait()
    if code != 0:
        raise RuntimeError(f"命令失败({code}): {' '.join(cmd)}\n{last}")
    _ops_update_step(task_id, step_idx, "done", "完成")

def _ops_zip_path(src_path, zip_path):
    os.makedirs(os.path.dirname(zip_path), exist_ok=True)
    if platform.system() == "Darwin" and os.path.isdir(src_path) and src_path.endswith(".app"):
        subprocess.run(
            ["ditto", "-c", "-k", "--sequesterRsrc", "--keepParent", src_path, zip_path],
            check=True,
        )
        return zip_path
    base = zip_path[:-4] if zip_path.lower().endswith(".zip") else zip_path
    if os.path.isdir(src_path):
        shutil.make_archive(base, "zip", root_dir=os.path.dirname(src_path), base_dir=os.path.basename(src_path))
    else:
        _ops_zip_files([{"path": src_path, "arcname": os.path.basename(src_path)}], zip_path)
    return zip_path if zip_path.lower().endswith(".zip") else base + ".zip"

def _ops_license_error_response():
    is_valid, msg = check_license()
    if is_valid:
        return None
    return jsonify({"status": "error", "message": msg or "授权未通过"}), 403

@ops_bp.route('/ops-tools')
def ops_tools_page():
    is_valid, _ = check_license()
    if not is_valid:
        return render_template('activate.html', hwid=CURRENT_HWID)
    return render_template('ops_tools.html')

@ops_bp.route('/ops-tools/raw')
def ops_tools_raw_page():
    is_valid, _ = check_license()
    if not is_valid:
        return render_template('activate.html', hwid=CURRENT_HWID)
    return render_template('ops_tools_raw.html')

@ops_bp.route('/api/ops/tasks/<task_id>/progress')
def api_ops_task_progress(task_id):
    license_err = _ops_license_error_response()
    if license_err:
        return license_err
    task = _ops_get_task(task_id)
    if not task:
        return jsonify({"status": "error", "message": "任务不存在"}), 404
    return jsonify({"status": "ok", "task": _ops_public_task(task)})

@ops_bp.route('/api/ops/tasks/<task_id>/download')
def api_ops_task_download(task_id):
    license_err = _ops_license_error_response()
    if license_err:
        return license_err
    task = _ops_get_task(task_id)
    if not task:
        return jsonify({"status": "error", "message": "任务不存在"}), 404
    path = task.get("result_path") or ""
    if task.get("status") != "done" or not path or not os.path.exists(path):
        return jsonify({"status": "error", "message": "结果文件尚未生成"}), 400
    resp = send_file(path, as_attachment=True, download_name=task.get("download_name") or os.path.basename(path))
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp

# License & Package Build routes moved to app_ops_extra.py
