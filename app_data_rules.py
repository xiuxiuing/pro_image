import os
import json
from flask import Blueprint, request, jsonify, send_file
from openpyxl import load_workbook

rules_bp = Blueprint('data_rules', __name__)


def init_rules(context):
    global dm, _validate_upload, DEFAULT_RULE_CATEGORIES_XLSX, CATEGORY_L1_BUCKET_TAGS_JSON
    dm = context["dm"]
    _validate_upload = context["validate_upload"]
    DEFAULT_RULE_CATEGORIES_XLSX = context["default_rule_categories_xlsx"]
    CATEGORY_L1_BUCKET_TAGS_JSON = context["category_l1_bucket_tags_json"]


def _norm_cell(v):
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in ("nan", "none", "null") else s

def _build_category_tree(rows):
    l1_aliases = ("美团一级类目", "一级类目", "美团类目一级", "美团1级类目")
    l2_aliases = ("美团二级类目", "二级类目", "美团类目二级", "美团2级类目")
    l3_aliases = ("美团三级类目", "三级类目", "美团类目三级", "美团3级类目")

    def first_val(item, aliases):
        for key in aliases:
            if key in item:
                v = _norm_cell(item.get(key))
                if v:
                    return v
        return ""

    l1_map = {}
    for row in rows:
        l3 = first_val(row, l3_aliases)
        if not l3:
            continue
        l1 = first_val(row, l1_aliases) or "未分类一级类目"
        l2 = first_val(row, l2_aliases) or "未分类二级类目"
        l1_entry = l1_map.setdefault(l1, {"name": l1, "children_map": {}, "l3_count": 0})
        l2_entry = l1_entry["children_map"].setdefault(l2, {"name": l2, "children": [], "seen": set()})
        if l3 not in l2_entry["seen"]:
            l2_entry["children"].append({"name": l3})
            l2_entry["seen"].add(l3)
            l1_entry["l3_count"] += 1

    items = []
    for l1_name in sorted(l1_map.keys()):
        l1_entry = l1_map[l1_name]
        children = []
        for l2_name in sorted(l1_entry["children_map"].keys()):
            l2_entry = l1_entry["children_map"][l2_name]
            children.append({
                "name": l2_name,
                "l3_count": len(l2_entry["children"]),
                "children": sorted(l2_entry["children"], key=lambda x: x["name"]),
            })
        items.append({
            "name": l1_name,
            "l3_count": l1_entry["l3_count"],
            "children": children,
        })

    return {
        "items": items,
        "l1_count": len(items),
        "l3_count": sum(item["l3_count"] for item in items),
    }

def _workbook_to_rows(wb):
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return []
    out = []
    for row in rows:
        if all(cell is None for cell in row):
            continue
        out.append(dict(zip(headers, row)))
    return out

def _excel_file_to_rows(file_storage):
    wb = load_workbook(file_storage, data_only=True)
    return _workbook_to_rows(wb)

def _excel_path_to_rows(path: str):
    wb = load_workbook(path, data_only=True)
    return _workbook_to_rows(wb)

def _load_bucket_tags():
    if not CATEGORY_L1_BUCKET_TAGS_JSON or not os.path.isfile(CATEGORY_L1_BUCKET_TAGS_JSON):
        return []
    try:
        with open(CATEGORY_L1_BUCKET_TAGS_JSON, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        return []
    tags = data.get("tags") if isinstance(data, dict) else []
    if not isinstance(tags, list):
        return []
    out = []
    for item in tags:
        if not isinstance(item, dict):
            continue
        out.append({
            "id": str(item.get("id") or "").strip(),
            "label": str(item.get("label") or item.get("id") or "").strip(),
            "l1": [str(v).strip() for v in (item.get("l1") or []) if str(v).strip()],
        })
    return out

@rules_bp.route("/api/rule-templates", methods=["GET", "POST"])
def api_rule_templates():
    if request.method == "GET":
        return jsonify({"status": "ok", "items": dm.list_rule_templates()})
    d = request.get_json(silent=True) or {}
    name = (d.get("name") or "").strip()
    if not name:
        return jsonify({"status": "error", "message": "模板名称不能为空"}), 400
    desc = (d.get("description") or "").strip()
    config = d.get("config")
    if not isinstance(config, dict):
        return jsonify({"status": "error", "message": "config 须为对象"}), 400
    rid = dm.create_rule_template(name, desc, config)
    return jsonify({"status": "ok", "id": rid})

@rules_bp.route("/api/rule-templates/<int:tid>", methods=["GET", "PUT", "DELETE"])
def api_rule_template_one(tid):
    if request.method == "GET":
        t = dm.get_rule_template(tid)
        if not t:
            return jsonify({"status": "error", "message": "不存在"}), 404
        return jsonify({"status": "ok", "item": t})
    if request.method == "DELETE":
        ok, err = dm.delete_rule_template(tid)
        if not ok:
            return jsonify({"status": "error", "message": err or "删除失败"}), 400
        return jsonify({"status": "ok"})
    d = request.get_json(silent=True) or {}
    name = (d.get("name") or "").strip()
    if not name:
        return jsonify({"status": "error", "message": "模板名称不能为空"}), 400
    desc = (d.get("description") or "").strip()
    config = d.get("config")
    if not isinstance(config, dict):
        return jsonify({"status": "error", "message": "config 须为对象"}), 400
    if not dm.update_rule_template(tid, name, desc, config):
        return jsonify({"status": "error", "message": "更新失败"}), 400
    return jsonify({"status": "ok"})

@rules_bp.route("/api/rule-category-template")
def api_rule_category_template():
    from flask import send_file
    import io
    from openpyxl import Workbook
    if os.path.isfile(DEFAULT_RULE_CATEGORIES_XLSX):
        return send_file(
            DEFAULT_RULE_CATEGORIES_XLSX,
            as_attachment=True,
            download_name="类目配置模板.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    wb = Workbook()
    ws = wb.active
    ws.title = "类目模板"
    ws.append(["美团一级类目", "美团二级类目", "美团三级类目"])
    ws.append(["饮料", "碳酸饮料", "可乐"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="类目配置模板.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@rules_bp.route("/api/rule-categories/parse", methods=["POST"])
def api_rule_categories_parse():
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"status": "error", "message": "请先上传类目文件"}), 400
    err = _validate_upload(file, "类目文件")
    if err:
        return jsonify({"status": "error", "message": err}), 400
    try:
        rows = _excel_file_to_rows(file)
        tree = _build_category_tree(rows)
    except Exception as exc:
        return jsonify({"status": "error", "message": f"类目文件解析失败：{exc}"}), 400
    if not tree["items"]:
        return jsonify({"status": "error", "message": "未解析到有效的三级类目，请检查表头是否包含一级/二级/三级类目字段"}), 400
    return jsonify({"status": "ok", "tree": tree})

@rules_bp.route("/api/rule-categories/default")
def api_rule_categories_default():
    if not DEFAULT_RULE_CATEGORIES_XLSX or not os.path.isfile(DEFAULT_RULE_CATEGORIES_XLSX):
        return jsonify({"status": "error", "message": "默认类目模板不存在"}), 404
    try:
        rows = _excel_path_to_rows(DEFAULT_RULE_CATEGORIES_XLSX)
        tree = _build_category_tree(rows)
    except Exception as exc:
        return jsonify({"status": "error", "message": f"默认类目解析失败：{exc}"}), 500
    return jsonify({"status": "ok", "tree": tree})

@rules_bp.route("/api/rule-categories/bucket-tags")
def api_rule_categories_bucket_tags():
    return jsonify({"status": "ok", "tags": _load_bucket_tags()})
