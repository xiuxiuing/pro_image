from openpyxl import load_workbook
from openpyxl import Workbook


def optimize_numeric_value(val):
    """
    Standardizes numeric values for internal processing and Excel export:
    1. Handles very large numbers as strings (IDs, barcodes >= 10^12)
    2. Converts whole-number floats to integers.
    3. Rounds other floats to 4 decimal places.
    """
    if isinstance(val, (int, float)):
        abs_val = abs(val)
        if abs_val >= 10**12:
            # Large IDs MUST be strings to avoid scientific notation and precision loss
            return str(int(val))
        elif isinstance(val, float):
            if val == int(val):
                return int(val)
            else:
                return round(val, 4)
    return val


def excel_to_list_dict(file_path, sheet_name=None):
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows = ws.iter_rows(values_only=True)

    # 第一行作为 key
    headers = next(rows)
    if not headers:
        return []

    data = []
    for row in rows:
        # 跳过空行
        if all(cell is None for cell in row):
            continue
        
        row_list = [optimize_numeric_value(cell) for cell in row]
        data.append(dict(zip(headers, row_list)))

    return data


def write_dict_list_to_excel(data, file_path="output_text.xlsx"):
    if not data:
        print("Empty data, nothing to write")
        return

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 写入表头
    headers = list(data[0].keys())
    ws.append(headers)

    # 写入每一行
    for item in data:
        row = [optimize_numeric_value(item.get(h, "")) for h in headers]
        ws.append(row)

    # 保存文件
    wb.save(file_path)
    print(f"Excel saved to {file_path}")


def get_sku_id(item):
    """
    Consistent SKU ID extraction from a dictionary item.
    Handles 'skuid', 'SKUID', and numeric string conversions.
    """
    keys = ["skuid", "SKUID", "skuId"]
    val = None
    for k in keys:
        if k in item and item[k] is not None:
            val = item[k]
            break
    
    if val is None or str(val).strip() == "":
        return ""
        
    s_val = str(val).strip()
    if s_val.endswith(".0"):
        return s_val[:-2]
    return s_val


def write_multisheet_dict_to_excel(sheet_data_dict, file_path="output_multisheet.xlsx"):
    """
    Writes a dictionary of { "Sheet Name": [dict, dict, ...] } to an Excel file with multiple sheets.
    """
    if not sheet_data_dict:
        print("Empty sheet data, nothing to write")
        return

    # Create workbook
    wb = Workbook()
    
    # Remove default sheets
    for sname in ["Sheet", "Sheet1"]:
        if sname in wb.sheetnames:
            wb.remove(wb[sname])

    for sheet_title, data in sheet_data_dict.items():
        ws = wb.create_sheet(title=sheet_title)
        
        if not data:
            continue
            
        # Write headers
        headers = list(data[0].keys())
        ws.append(headers)

        # Write rows
        for item in data:
            row = [optimize_numeric_value(item.get(h, "")) for h in headers]
            ws.append(row)

    # Save file
    wb.save(file_path)
    print(f"Multi-sheet Excel saved to {file_path}")
