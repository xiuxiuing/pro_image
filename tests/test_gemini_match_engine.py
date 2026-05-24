import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


def make_workbook(path, headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)


def read_rows(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    return [dict(zip(headers, row)) for row in rows[1:]]


class FakeGeminiMatcher:
    def __init__(self):
        self.calls = []

    def match_one(self, main_item, store_id, candidates):
        self.calls.append((main_item, store_id, list(candidates)))
        main_name = main_item["name"]
        for candidate in candidates:
            if candidate["name"] == main_name and candidate["spec"] == main_item["spec"]:
                return {
                    "matched": True,
                    "comp_sku_id": candidate["sku_id"],
                    "confidence": 0.96,
                    "reason": "same product and same spec",
                    "spec_check": "same",
                    "reject_reason": "",
                }
        return {
            "matched": False,
            "comp_sku_id": "",
            "confidence": 0.0,
            "reason": "no same product with same spec",
            "spec_check": "different",
            "reject_reason": "no exact match",
        }


class GeminiMatchEngineTests(unittest.TestCase):
    def test_default_file_discovery_uses_main_and_comp_files_only(self):
        from gemini_match_engine.defaults import discover_default_files

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            headers = ["skuid", "商品名称", "规格"]
            make_workbook(tmp / "竞店D.xlsx", headers, [])
            make_workbook(tmp / "竞店B.xlsx", headers, [])
            make_workbook(tmp / "竞店C.xlsx", headers, [])
            make_workbook(tmp / "主店.xlsx", headers, [])
            make_workbook(tmp / "output_主店与竞店匹配结果.xlsx", headers, [])
            make_workbook(tmp / "竞店A.xlsx", headers, [])

            main_path, comp_paths = discover_default_files(str(tmp))

            self.assertEqual(Path(main_path).name, "主店.xlsx")
            self.assertEqual([Path(p).name for p in comp_paths], ["竞店A.xlsx", "竞店B.xlsx", "竞店C.xlsx", "竞店D.xlsx"])

    def test_outputs_existing_format_and_leaves_unmatched_blank(self):
        from gemini_match_engine.engine import run_gemini_match_analysis

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            headers = [
                "skuid", "商品名称", "规格", "图片", "条码",
                "一级类目", "二级类目", "三级类目", "活动价", "原价", "销量",
            ]
            main_path = tmp / "main.xlsx"
            comp_path = tmp / "comp.xlsx"
            make_workbook(
                main_path,
                headers,
                [
                    ["m1", "可乐", "500ml", "", "6901", "饮品", "饮料", "碳酸饮料", 3, 4, 10],
                    ["m2", "雪碧", "330ml", "", "6902", "饮品", "饮料", "碳酸饮料", 2, 3, 8],
                ],
            )
            make_workbook(
                comp_path,
                headers,
                [
                    ["c1", "可乐", "500ml", "", "7901", "饮品", "饮料", "碳酸饮料", 2.8, 4, 20],
                    ["c2", "雪碧", "500ml", "", "7902", "饮品", "饮料", "碳酸饮料", 3.5, 5, 6],
                ],
            )

            out_path = run_gemini_match_analysis(
                str(main_path),
                [str(comp_path)],
                output_name="gemini_test",
                output_dir=str(tmp),
                matcher=FakeGeminiMatcher(),
            )

            rows = read_rows(out_path)
            self.assertEqual(rows[0]["skuId"], "m1")
            self.assertEqual(rows[0]["0skuId"], "c1")
            self.assertEqual(rows[0]["0匹配"], "Gemini匹配")
            self.assertAlmostEqual(float(rows[0]["0相似度"]), 0.96)
            self.assertEqual(rows[1]["skuId"], "m2")
            self.assertIn(rows[1].get("0skuId"), (None, ""))
            self.assertIn(rows[1].get("0匹配"), (None, ""))

    def test_shards_large_candidate_sets_without_business_filtering(self):
        from gemini_match_engine.engine import run_gemini_match_analysis

        class LastShardMatcher(FakeGeminiMatcher):
            def match_one(self, main_item, store_id, candidates):
                self.calls.append((main_item, store_id, list(candidates)))
                for candidate in candidates:
                    if candidate["sku_id"] == "c7":
                        return {
                            "matched": True,
                            "comp_sku_id": "c7",
                            "confidence": 0.91,
                            "reason": "selected from shard",
                            "spec_check": "same",
                            "reject_reason": "",
                        }
                return {
                    "matched": False,
                    "comp_sku_id": "",
                    "confidence": 0.0,
                    "reason": "",
                    "spec_check": "",
                    "reject_reason": "not in this shard",
                }

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            headers = ["skuid", "商品名称", "规格", "图片", "条码", "一级类目", "二级类目", "三级类目", "活动价", "原价", "销量"]
            main_path = tmp / "main.xlsx"
            comp_path = tmp / "comp.xlsx"
            make_workbook(main_path, headers, [["m1", "目标商品", "1个", "", "", "日用", "日用", "日用", 10, 12, 1]])
            make_workbook(
                comp_path,
                headers,
                [[f"c{i}", f"候选{i}", "1个", "", "", "不同", "不同", "不同", 1, 2, i] for i in range(8)],
            )

            matcher = LastShardMatcher()
            out_path = run_gemini_match_analysis(
                str(main_path),
                [str(comp_path)],
                output_name="gemini_shard",
                output_dir=str(tmp),
                matcher=matcher,
                shard_size=3,
            )

            rows = read_rows(out_path)
            self.assertEqual(rows[0]["0skuId"], "c7")
            self.assertEqual(len(matcher.calls), 4)
            self.assertEqual([len(call[2]) for call in matcher.calls[:3]], [3, 3, 2])
            self.assertEqual([candidate["sku_id"] for candidate in matcher.calls[3][2]], ["c7"])

    def test_limits_main_rows_and_prefers_relevant_candidates_for_gemini(self):
        from gemini_match_engine.engine import run_gemini_match_analysis

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            headers = ["skuid", "商品名称", "规格", "图片", "条码", "一级类目", "二级类目", "三级类目", "活动价", "原价", "销量"]
            main_path = tmp / "main.xlsx"
            comp_path = tmp / "comp.xlsx"
            make_workbook(
                main_path,
                headers,
                [
                    ["m1", "蓝月亮洗衣液", "500g", "", "", "清洁", "洗衣", "洗衣液", 10, 12, 1],
                    ["m2", "第二个商品", "1个", "", "", "日用", "日用", "日用", 10, 12, 1],
                ],
            )
            make_workbook(
                comp_path,
                headers,
                [
                    ["c1", "无关商品1", "1个", "", "", "其他", "其他", "其他", 1, 2, 1],
                    ["c2", "无关商品2", "1个", "", "", "其他", "其他", "其他", 1, 2, 1],
                    ["c3", "蓝月亮洗衣液", "500g", "", "", "不同", "不同", "不同", 1, 2, 1],
                    ["c4", "无关商品3", "1个", "", "", "其他", "其他", "其他", 1, 2, 1],
                ],
            )

            matcher = FakeGeminiMatcher()
            out_path = run_gemini_match_analysis(
                str(main_path),
                [str(comp_path)],
                output_name="gemini_limited",
                output_dir=str(tmp),
                matcher=matcher,
                max_main_rows=1,
                candidate_limit=2,
            )

            rows = read_rows(out_path)
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["skuId"], "m1")
            self.assertEqual(rows[0]["0skuId"], "c3")
            self.assertEqual(len(matcher.calls[0][2]), 2)
            self.assertIn("c3", [candidate["sku_id"] for candidate in matcher.calls[0][2]])


if __name__ == "__main__":
    unittest.main()
