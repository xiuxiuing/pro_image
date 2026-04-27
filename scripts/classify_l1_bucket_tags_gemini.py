# -*- coding: utf-8 -*-
"""
根据 default_meituan_categories.xlsx（或其它同结构表）中的一级/二/三级类目名称，
用 Gemini 为每个**一级类目**在「固定 8 标签」中归纳分类，并写出 match_rule 页用的
data/category_l1_bucket_tags.json。

固定标签（须与业务约定一致，改标签请改脚本内 FIXED_BUCKETS 与 data 文件）：
  休食快消、服饰箱包、3C数码、家居百货、个护家清、生鲜果蔬、成人用品、化妆收纳

环境变量: GEMINI_API_KEY（与 extract_info_ai2 相同）

用法:
  # 只根据内置启发式表写 JSON（不调用 Gemini，离线可用）
  python3 scripts/classify_l1_bucket_tags_gemini.py --heuristic-only

  # 调 Gemini 全量重算（会读入 xlsx，构造提示词，写 JSON）
  export GEMINI_API_KEY="..."
  python3 scripts/classify_l1_bucket_tags_gemini.py \\
    --input data/default_meituan_categories.xlsx \\
    -o data/category_l1_bucket_tags.json

  # 只打印将发给模型的 prompt，不请求 API
  python3 scripts/classify_l1_bucket_tags_gemini.py --dry-run
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from collections import defaultdict
from typing import Any, Dict, List, Tuple

# 项目根
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from openpyxl import load_workbook
from google import genai
from pydantic import BaseModel, Field

try:
    from extract_info_ai2 import DEFAULT_MODEL_NAME
except Exception:
    DEFAULT_MODEL_NAME = "models/gemini-3.1-flash-lite-preview"

# 固定 8 标签：id -> 展示名（与前端/配置一致）
FIXED_BUCKETS: List[Tuple[str, str]] = [
    ("leisure_fmcg", "休食快消"),
    ("fashion_bags", "服饰箱包"),
    ("digital_3c", "3C数码"),
    ("home_general", "家居百货"),
    ("care_clean", "个护家清"),
    ("fresh", "生鲜果蔬"),
    ("adult", "成人用品"),
    ("makeup_storage", "化妆收纳"),
]
ALLOWED = {b for _, b in FIXED_BUCKETS}


def _heuristic_l1_to_bucket(l1: str) -> str:
    """
    无 Gemini 时的兜底划分（可随你人工调整，最终以 Gemini 生成为准时会被覆盖）。
    """
    s = (l1 or "").strip()
    m = {
        "休食快消": [
            "休闲食品", "乳品", "营养冲调", "粮油调味干货", "速食/罐头", "酒类", "雪糕/冰淇淋/食用冰", "饮品",
        ],
        "服饰箱包": ["服饰鞋包", "运动户外", "珠宝首饰", "手表眼镜"],
        "3C数码": ["手机通讯", "电脑数码", "家用电器", "汽车用品"],
        "家居百货": [
            "厨具餐具", "学习/办公用品", "宠物生活", "家居日用", "家纺布艺", "家装建材", "店铺管理",
            "母婴用品", "玩具乐器", "节庆礼品", "花卉园艺",
        ],
        "个护家清": ["个人洗护", "家庭清洁", "医疗器械"],
        "生鲜果蔬": ["水果", "蔬菜/豆制品", "生肉/生禽/生蛋", "熟食/鲜食", "速冻食品"],
        "成人用品": ["成人用品"],
        "化妆收纳": ["彩妆香水", "美容护肤"],
    }
    for b, l1s in m.items():
        if s in l1s:
            return b
    return "家居百货"


def _read_l1_rows(path: str) -> List[Dict[str, str]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(c or "").strip() for c in rows[0]]
    def _col(aliases, fallback_idx=None):
        for a in aliases:
            if a in headers:
                return headers.index(a)
        return fallback_idx
    i1 = _col(["美团一级类目", "一级类目", "美团类目一级"])
    i2 = _col(["美团二级类目", "二级类目", "美团类目二级"])
    i3 = _col(["美团三级类目", "三级类目", "美团类目三级"])
    if i1 is None or i2 is None or i3 is None:
        raise SystemExit("表头需包含 美团一/二/三级类目 列名")
    out = []
    for row in rows[1:]:
        if not row or all(c is None for c in row):
            continue
        d = {headers[j]: row[j] for j in range(min(len(headers), len(row))) if j < len(headers)}
        l1, l2, l3 = d.get(headers[i1]), d.get(headers[i2]), d.get(headers[i3])
        l1s = re.sub(r"\s+", " ", str(l1 or "").strip())
        l2s = re.sub(r"\s+", " ", str(l2 or "").strip())
        l3s = re.sub(r"\s+", " ", str(l3 or "").strip())
        if l3s and l3s.lower() not in ("nan", "none"):
            out.append({"l1": l1s, "l2": l2s, "l3": l3s})
    return out


def _aggregate_l1_samples(
    rows: List[Dict[str, str]], max_l2: int = 5, l3_per_l2: int = 3
) -> List[Dict[str, Any]]:
    """每个一级下抽若干 二级 → 三级 样例，供模型理解。"""

    l1_to_pairs: Dict[str, List[Tuple[str, str]]] = defaultdict(list)
    for r in rows:
        l1 = (r.get("l1") or "未分类一级").strip() or "未分类一级"
        l2 = (r.get("l2") or "").strip() or "未分类二级"
        l3 = (r.get("l3") or "").strip()
        if not l3:
            continue
        l1_to_pairs[l1].append((l2, l3))

    result: List[Dict[str, Any]] = []
    for l1 in sorted(l1_to_pairs.keys()):
        pairs = l1_to_pairs[l1]
        l2_order = list(dict.fromkeys(p[0] for p in pairs))[:max_l2]
        samples: List[Dict[str, Any]] = []
        for l2k in l2_order:
            l3s = [p[1] for p in pairs if p[0] == l2k]
            l3s = list(dict.fromkeys(l3s))[:l3_per_l2]
            if l3s:
                samples.append({"二级": l2k, "三级示例": l3s})
        if not samples and pairs:
            samples.append({"二级": pairs[0][0], "三级示例": [pairs[0][1]]})
        result.append({"美团一级类目": l1, "样例_二级到三级": samples})
    return result


class OneAssignment(BaseModel):
    l1: str = Field(description="美团一级类目原文字符串，须与输入一致")
    bucket: str = Field(description="8 个固定标签之一，逐字与枚举一致")


class BatchAssign(BaseModel):
    items: List[OneAssignment]


def _build_prompt(aggregated: List[Dict[str, Any]]) -> str:
    lines = [f"- {b}" for _id, b in FIXED_BUCKETS]
    enum_text = "\n".join(f"{i+1}. {b}" for i, b in enumerate(b for _a, b in FIXED_BUCKETS))
    data_json = json.dumps(aggregated, ensure_ascii=False, indent=2)
    return f"""你是中国即时零售（美团 O2O）平台的类目与商品运营专家。

下面 JSON 的每一项是一个「美团一级类目」名称，并附带该一级下少量「二级、三级」名称样例，**仅作理解含义之用**（二三级不必穷尽）。

请为**每一个**一级类目，在下列 **8 个固定标签** 中**只选其一**、且标签文字必须与列表**逐字相同**（不可新增、不可改写、不可合并）：

{enum_text}

划分原则（在边界不清时取「更贴近消费场景」的为主）：
- 休食快消：包装类零食、水饮、乳饮冲调、油盐酱醋、即食罐头、酒、冰淇淋等**非现制生鲜**的包装食品。
- 服饰箱包：鞋服、包、表镜配饰、运动户外等可穿戴/可提拿穿戴。
- 3C数码：手机、电脑、小家电/大家电、汽车电子用品等**带电弱电强电商品**；店铺工具偏「设备」也可归此（若你更确定归家居百货也可说明理由，但只能选一个标签）。
- 家居百货：家清以外的家居生活、厨房用具、床品、建材园艺、礼赠节庆、文教具、婴童耐用品/玩具、宠用品（非主粮场景偏百货时）等**生活方式类**。
- 个护家清：纸品湿巾、家清、个人洗护、健康器械/防护等**偏清洁与健康护理非彩妆**；若与化妆收纳二选一有歧义，以「非彩妆上脸」为个护家清。
- 生鲜果蔬：生鲜、蔬果豆品、肉禽蛋、现制/短保熟食、冻品等**易腐鲜食**。
- 成人用品：仅限成人用品一级本身或明确为该类。
- 化妆收纳：彩妆、护肤、与化妆护肤强相关的**收纳/工具**（若一级本身以美妆为主）。

请输出**严格仅 JSON**（不要 markdown 代码块），且符合给定 schema。必须覆盖**全部**输入的一级类目，**每个 l1 恰好一条**，`bucket` 必须是上述 8 个之一。

【输入数据】
{data_json}
"""


def _call_gemini(prompt: str, api_key: str, model: str) -> BatchAssign:
    client = genai.Client(api_key=api_key)
    for attempt in range(5):
        try:
            resp = client.models.generate_content(
                model=model,
                contents=prompt,
                config={"response_mime_type": "application/json", "response_schema": BatchAssign},
            )
            if resp.parsed and isinstance(resp.parsed, BatchAssign):
                return resp.parsed
        except Exception as e:
            msg = str(e)
            if "429" in msg or "RESOURCE_EXHAUSTED" in msg:
                time.sleep((attempt + 1) * 25)
            else:
                time.sleep(5)
    raise RuntimeError("Gemini 未返回可解析的 BatchAssign")


def _assignments_to_tags(assignments: List[OneAssignment]) -> List[Dict[str, Any]]:
    by_bucket: Dict[str, List[str]] = defaultdict(list)
    for a in assignments:
        b = a.bucket.strip()
        if b not in ALLOWED:
            raise SystemExit(f"非法 bucket: {b!r}，仅允许: {sorted(ALLOWED)}")
        if a.l1 not in by_bucket[b]:
            by_bucket[b].append(a.l1)

    tags = []
    for bid, blabel in FIXED_BUCKETS:
        l1s = sorted(by_bucket.get(blabel, []))
        tags.append({"id": bid, "label": blabel, "l1": l1s})
    return tags


def _heuristic_to_tags(l1_list: List[str]) -> List[Dict[str, Any]]:
    by_bucket: Dict[str, List[str]] = defaultdict(list)
    for l1 in l1_list:
        b = _heuristic_l1_to_bucket(l1)
        by_bucket[b].append(l1)
    tags = []
    for bid, blabel in FIXED_BUCKETS:
        l1s = sorted(set(by_bucket.get(blabel, [])))
        tags.append({"id": bid, "label": blabel, "l1": l1s})
    return tags


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", default=os.path.join(ROOT, "data", "default_meituan_categories.xlsx"))
    ap.add_argument("-o", "--output", default=os.path.join(ROOT, "data", "category_l1_bucket_tags.json"))
    ap.add_argument("--heuristic-only", action="store_true", help="不调用 API，用内置启发式表写 -o")
    ap.add_argument("--dry-run", action="store_true", help="只打印将发送的提示词，不写文件")
    ap.add_argument("--model", default=DEFAULT_MODEL_NAME)
    args = ap.parse_args()

    rows = _read_l1_rows(args.input)
    if not rows:
        raise SystemExit("未读到有效行，请检查 xlsx 路径与表头。")

    l1_unique = sorted({r["l1"] for r in rows if r.get("l1")})
    ag = _aggregate_l1_samples(rows)
    prompt = _build_prompt(ag)

    if args.dry_run:
        print(prompt)
        return

    if args.heuristic_only:
        out = {
            "version": 1,
            "source": "heuristic",
            "tags": _heuristic_to_tags(l1_unique),
        }
        os.makedirs(os.path.dirname(args.output) or ".", exist_ok=True)
        with open(args.output, "w", encoding="utf-8") as f:
            json.dump(out, f, ensure_ascii=False, indent=2)
        print(f"已写入 {args.output}（heuristic 共 {len(l1_unique)} 个一级）")
        return

    key = (os.environ.get("GEMINI_API_KEY") or "").strip()
    if not key:
        raise SystemExit("未设置 GEMINI_API_KEY，或加 --heuristic-only / --dry-run")

    parsed = _call_gemini(prompt, key, args.model.strip() or DEFAULT_MODEL_NAME)
    got = {a.l1.strip() for a in parsed.items if a.l1}
    missing = set(l1_unique) - got
    extra = got - set(l1_unique)
    if missing:
        raise SystemExit(f"模型漏掉一级类目: {sorted(missing)[:10]}{'...' if len(missing) > 10 else ''}")
    if extra:
        # 松一点：可警告
        print("警告: 模型多出一级类目", sorted(extra)[:5])

    out = {
        "version": 1,
        "source": "gemini",
        "model": args.model,
        "tags": _assignments_to_tags(list(parsed.items)),
    }
    os.makedirs(os.path.dirname(args.output) or ".", exist_ok=True)
    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print(f"已写入 {args.output}（gemini, {len(parsed.items)} 条）")


if __name__ == "__main__":
    main()
