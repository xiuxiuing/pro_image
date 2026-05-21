import os
import re
import time
import zipfile
import shutil
import tempfile
import pandas as pd
import utils
from openpyxl import Workbook
from data_mgr_base import INTERNAL_EXPORT_KEYS

class DataManagerExportMixin:
    def _export_safe_filename(self, name, fallback):
        text = str(name or fallback).strip() or fallback
        text = re.sub(r'[\\/:*?"<>|\r\n]+', "_", text)
        return text[:80] or fallback

    def _export_main_operation_note(self, row, store_map):
        notes = []
        if str(row.get("是否淘汰", "")).strip() == "是" or str(row.get("淘汰标记", "")).strip() == "1":
            notes.append("标记淘汰")
        if str(row.get("is_handled", "")).strip() == "1":
            notes.append("已处理")
        follow_store = str(row.get("跟价店", "")).strip()
        if follow_store:
            notes.append(f"跟价店：{follow_store}")
        if str(row.get("新售价", "")).strip() or str(row.get("新活动价", "")).strip():
            notes.append("修改价格")
        ref_name_store = str(row.get("ref_name_store", "")).strip()
        if ref_name_store:
            notes.append(f"名称参考：{store_map.get(ref_name_store, ref_name_store)}")
        ref_image_store = str(row.get("ref_image_store", "")).strip()
        if ref_image_store:
            notes.append(f"图片参考：{store_map.get(ref_image_store, ref_image_store)}")
        return "；".join(notes)

    def _export_comp_operation_note(self, row):
        notes = []
        if str(row.get("关联主店skuId", "")).strip():
            notes.append("已关联")
        else:
            notes.append("未关联")
        match_type = str(row.get("关联方式", "")).strip()
        if match_type:
            notes.append(match_type)
        if str(row.get("是否新增", "")).strip() == "是":
            notes.append("标记新增")
        if str(row.get("是否不处理", "")).strip() == "是":
            notes.append("不处理")
        return "；".join(notes)

    def _blank_if_missing(self, row, col):
        val = row.get(col, "")
        if pd.isna(val):
            return ""
        text = str(val).strip()
        return "" if text.lower() in ("nan", "none") else text

    def _build_platform_price_sheets(self, main_change_df):
        records = main_change_df.fillna("").to_dict(orient="records") if main_change_df is not None else []

        qnh_rows = []
        meituan_rows = []
        eleme_rows = []
        for row in records:
            sku = self._blank_if_missing(row, "skuId")
            name = self._blank_if_missing(row, "商品名称")
            new_activity_price = self._blank_if_missing(row, "新活动价")
            new_retail_price = self._blank_if_missing(row, "新售价")

            qnh_rows.append({
                "*商品规格（SKUID)": sku,
                "*门店编码": "",
                "*商品零售价（元）": new_retail_price,
                "渠道编码": "",
            })
            meituan_rows.append({
                "UPC（条形码）": "",
                "SKU码/货号": sku,
                "商品名称": name,
                "活动价": new_activity_price,
                "折扣率": "",
                "每单限购份数": "1",
                "当日活动库存": "-1",
            })
            eleme_rows.append({
                "商品条形码": "",
                "自定义ID": sku,
                "活动价": new_activity_price,
                "活动总库存": "",
                "每日活动库存": "",
                "每人/活动期间限购": "1",
                "每人/每日限购": "1",
            })
        return qnh_rows, meituan_rows, eleme_rows

    def _write_export_new_workbook(self, sheet_data, path):
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)

        def add_records_sheet(title, records, headers=None):
            ws = wb.create_sheet(title=title)
            headers = list(headers or (records[0].keys() if records else []))
            if headers:
                ws.append(headers)
                for item in records:
                    ws.append([utils.optimize_numeric_value(item.get(h, "")) for h in headers])
            return ws

        add_records_sheet("新增(竞店)", sheet_data.get("新增(竞店)", []))
        add_records_sheet("淘汰(主店)", sheet_data.get("淘汰(主店)", []))
        add_records_sheet(
            "牵牛花零售价修改",
            sheet_data.get("牵牛花零售价修改", []),
            ["*商品规格（SKUID)", "*门店编码", "*商品零售价（元）", "渠道编码"],
        )
        add_records_sheet(
            "美团商家版-活动价修改第二页",
            sheet_data.get("美团商家版-活动价修改第二页", []),
            ["UPC（条形码）", "SKU码/货号", "商品名称", "活动价", "折扣率", "每单限购份数", "当日活动库存"],
        )
        add_records_sheet(
            "饿了么商家版-活动价修改",
            sheet_data.get("饿了么商家版-活动价修改", []),
            ["商品条形码", "自定义ID", "活动价", "活动总库存", "每日活动库存", "每人/活动期间限购", "每人/每日限购"],
        )

        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    cell.number_format = "@"
            for col_cells in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
                ws.column_dimensions[col_cells[0].column_letter].width = min(max(max_len + 2, 12), 36)
        wb.save(path)

    def save_to_excel(self):
        filename = f"对比分析全量成果_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
        dirs = self._get_project_dirs(self.active_project_id)
        path = os.path.join(dirs["outputs"], filename)
        
        export_df = self.grid_df.copy()
        cols_to_drop = [c for c in INTERNAL_EXPORT_KEYS if c in export_df.columns]
        if cols_to_drop: export_df.drop(columns=cols_to_drop, inplace=True)
        utils.write_dict_list_to_excel(export_df.fillna("").to_dict(orient='records'), path)
        return path

    def save_separate_exports(self):
        temp_dir = tempfile.mkdtemp()
        try:
            store_id_to_name = {str(i): name for i, name in enumerate(self.store_names)}
            op_time = time.strftime('%Y-%m-%d %H:%M:%S')

            with self._db_lock, self._get_conn() as conn:
                main_export = pd.read_sql(
                    "SELECT * FROM main_products WHERE project_id = ?",
                    conn,
                    params=(self.active_project_id,),
                )
                if main_export.empty:
                    main_export = pd.DataFrame()
                else:
                    for col in ["是否淘汰", "淘汰标记", "新售价", "新活动价", "跟价店", "is_handled", "ref_name_store", "ref_image_store"]:
                        if col not in main_export.columns:
                            main_export[col] = ""
                    main_export["店铺"] = self.main_store_name or "主店"
                    main_export["操作记录"] = main_export.apply(lambda r: self._export_main_operation_note(r, store_id_to_name), axis=1)
                    main_export["导出时间"] = op_time
                    drop_cols = [c for c in ["project_id", "_row_orig_idx"] if c in main_export.columns]
                    if drop_cols:
                        main_export.drop(columns=drop_cols, inplace=True)
                    leading = [c for c in ["店铺", "skuId", "商品名称", "规格名称", "主图链接", "销售", "原价", "活动价", "采购价", "新售价", "新活动价", "跟价店", "是否淘汰", "is_handled", "操作记录", "导出时间"] if c in main_export.columns]
                    main_export = main_export[leading + [c for c in main_export.columns if c not in leading]]
                main_file = self._export_safe_filename(f"主店_{self.main_store_name}", "主店")
                main_export.fillna("").to_excel(os.path.join(temp_dir, f"{main_file}.xlsx"), index=False)

                for i, store_name in enumerate(self.store_names):
                    store_id = str(i)
                    comp_df = pd.read_sql(
                        "SELECT * FROM comp_products WHERE project_id = ? AND store_id = ?",
                        conn,
                        params=(self.active_project_id, store_id),
                    )
                    links_df = pd.read_sql(
                        """
                        SELECT
                            pl.main_sku_id AS 关联主店skuId,
                            pl.comp_sku_id,
                            pl.store_id,
                            pl.similarity AS 匹配相似度,
                            pl.match_type AS 关联方式,
                            pl.is_new_add AS 链接是否新增,
                            mp.`商品名称` AS 关联主店商品名称
                        FROM product_links pl
                        LEFT JOIN main_products mp
                          ON mp.project_id = pl.project_id
                         AND mp.skuId = pl.main_sku_id
                        WHERE pl.project_id = ? AND pl.store_id = ?
                        """,
                        conn,
                        params=(self.active_project_id, store_id),
                    )
                    if not links_df.empty:
                        links_df = links_df.drop_duplicates(subset=["store_id", "comp_sku_id"], keep="first")
                    if comp_df.empty:
                        comp_df = pd.DataFrame(columns=["store_id", "skuId"])
                    comp_export = comp_df.merge(
                        links_df,
                        left_on=["store_id", "skuId"],
                        right_on=["store_id", "comp_sku_id"],
                        how="left",
                    )
                    if "is_new_add" not in comp_export.columns:
                        comp_export["is_new_add"] = ""
                    if "is_ignored" not in comp_export.columns:
                        comp_export["is_ignored"] = ""
                    comp_export["是否新增"] = comp_export["is_new_add"].where(
                        comp_export["is_new_add"].fillna("").astype(str).str.strip() != "",
                        comp_export.get("链接是否新增", ""),
                    )
                    comp_export["是否不处理"] = comp_export["is_ignored"]
                    comp_export["店铺"] = store_name
                    comp_export["关联状态"] = comp_export["关联主店skuId"].fillna("").astype(str).str.strip().map(lambda v: "已关联" if v else "未关联")
                    comp_export["操作记录"] = comp_export.apply(self._export_comp_operation_note, axis=1)
                    comp_export["导出时间"] = op_time
                    drop_cols = [c for c in ["project_id", "store_id", "comp_sku_id", "链接是否新增", "is_new_add", "is_ignored"] if c in comp_export.columns]
                    if drop_cols:
                        comp_export.drop(columns=drop_cols, inplace=True)
                    leading = [c for c in ["店铺", "关联状态", "关联主店skuId", "关联主店商品名称", "关联方式", "匹配相似度", "是否新增", "是否不处理", "操作记录", "导出时间", "skuId", "商品名称", "规格名称", "主图链接", "销售", "原价", "活动价", "采购价"] if c in comp_export.columns]
                    comp_export = comp_export[leading + [c for c in comp_export.columns if c not in leading]]
                    comp_file = self._export_safe_filename(f"竞店_{store_name}", f"竞店_{store_id}")
                    comp_export.fillna("").to_excel(os.path.join(temp_dir, f"{comp_file}.xlsx"), index=False)
            
            dirs = self._get_project_dirs(self.active_project_id)
            zip_path = os.path.join(dirs["outputs"], f"对比成果_{time.strftime('%Y%m%d_%H%M%S')}.zip")
            with zipfile.ZipFile(zip_path, 'w') as zf:
                for root, _, files in os.walk(temp_dir):
                    for file in files: zf.write(os.path.join(root, file), arcname=file)
            return zip_path
        finally: shutil.rmtree(temp_dir)

    def export_new_items(self):
        op_time = time.strftime('%Y-%m-%d %H:%M:%S')
        self._calculate_margins()
        
        # Mapping prefix to store name
        store_map = {str(i): name for i, name in enumerate(self.store_names)}
        
        # 1. Fetch ALL products marked as "New" from comp_products (includes linked + unlinked)
        all_new_data = []
        with self._db_lock, self._get_conn() as conn:
            comp_cols = {
                row[1] for row in conn.execute("PRAGMA table_info(comp_products)").fetchall()
            }
            if "is_new_add" in comp_cols:
                ignored_clause = "AND COALESCE(is_ignored, '') != '是'" if "is_ignored" in comp_cols else ""
                query = f"SELECT * FROM comp_products WHERE project_id = ? AND is_new_add = '是' {ignored_clause}"
                all_comp_new_df = pd.read_sql(query, conn, params=(self.active_project_id,))
            else:
                all_comp_new_df = pd.DataFrame()
            if not all_comp_new_df.empty:
                # Add store name and main store link info if available
                all_comp_new_df['竞品店铺'] = all_comp_new_df['store_id'].map(store_map)
                
                # Fetch links to get Main Store SKU if linked
                link_query = "SELECT comp_sku_id, main_sku_id, store_id FROM product_links WHERE project_id = ?"
                links = pd.read_sql(link_query, conn, params=(self.active_project_id,))
                
                # Merge to get Main SKU (store-aware to avoid cross-store false joins)
                merged = all_comp_new_df.merge(
                    links,
                    left_on=['skuId', 'store_id'],
                    right_on=['comp_sku_id', 'store_id'],
                    how='left'
                )
                merged.rename(columns={'main_sku_id': '主店SKU'}, inplace=True)
                merged['来源'] = merged['主店SKU'].apply(lambda v: '已匹配' if str(v).strip() not in ['', 'nan', 'None'] else '未匹配池')
                
                # Ensure core columns
                cols = ['来源', '主店SKU', '竞品店铺', 'skuId', '主图链接', '商品名称', '规格名称', '活动价', '原价', '销售', '条码']
                final_new_df = merged[[c for c in cols if c in merged.columns]].copy()
                for c in cols:
                    if c not in final_new_df.columns: final_new_df[c] = ""
                all_new_data = final_new_df.fillna("").to_dict(orient='records')

        final_df = pd.DataFrame(all_new_data) if all_new_data else pd.DataFrame(columns=["主店SKU", "竞品店铺", "skuId", "主图链接", "商品名称", "规格名称", "活动价", "原价", "销售", "条码"])
        final_df["操作时间"] = op_time

        # 2. Main store: eliminated / price-matched / ref-marked items
        mask = pd.Series(False, index=self.grid_df.index)
        if '是否淘汰' in self.grid_df.columns: mask |= (self.grid_df['是否淘汰'] == "是")
        if '跟价店' in self.grid_df.columns: mask |= (self.grid_df['跟价店'].notna() & (self.grid_df['跟价店'] != ""))
        if 'ref_name_store' in self.grid_df.columns: mask |= (self.grid_df['ref_name_store'].fillna('') != '')
        if 'ref_image_store' in self.grid_df.columns: mask |= (self.grid_df['ref_image_store'].fillna('') != '')
            
        elim_df = self.grid_df[mask].copy()
        if not elim_df.empty:
            main_cols = [c for c in elim_df.columns if (not c or not c[0].isdigit()) and c not in INTERNAL_EXPORT_KEYS]
            elim_export = elim_df[main_cols].copy()
            ref_name_s = elim_df['ref_name_store'].fillna('') if 'ref_name_store' in elim_df.columns else pd.Series('', index=elim_df.index)
            ref_image_s = elim_df['ref_image_store'].fillna('') if 'ref_image_store' in elim_df.columns else pd.Series('', index=elim_df.index)
            elim_export['名称参考店铺'] = ref_name_s.map(lambda v: store_map.get(str(v), '') if v else '')
            elim_export['参考商品名称'] = [
                elim_df.at[i, str(v) + '商品名称'] if v and (str(v) + '商品名称') in elim_df.columns else ''
                for i, v in ref_name_s.items()
            ]
            elim_export['图片参考店铺'] = ref_image_s.map(lambda v: store_map.get(str(v), '') if v else '')
            elim_export['参考图片链接'] = [
                elim_df.at[i, str(v) + '主图链接'] if v and (str(v) + '主图链接') in elim_df.columns else ''
                for i, v in ref_image_s.items()
            ]
            elim_export["操作时间"] = op_time
            elim_df = elim_export
        else: elim_df = pd.DataFrame(columns=["skuId", "主图链接", "商品名称", "规格名称", "活动价", "原价", "销售", "条码", "操作时间"])

        dirs = self._get_project_dirs(self.active_project_id)
        path = os.path.join(dirs["outputs"], f"新增竞品数据_{time.strftime('%Y%m%d_%H%M%S')}.xlsx")
        qnh_rows, meituan_rows, eleme_rows = self._build_platform_price_sheets(elim_df)
        sheet_data = {
            "新增(竞店)": final_df.fillna("").to_dict(orient='records'),
            "淘汰(主店)": elim_df.fillna("").to_dict(orient='records'),
            "牵牛花零售价修改": qnh_rows,
            "美团商家版-活动价修改第二页": meituan_rows,
            "饿了么商家版-活动价修改": eleme_rows,
        }
        self._write_export_new_workbook(sheet_data, path)
        return path
